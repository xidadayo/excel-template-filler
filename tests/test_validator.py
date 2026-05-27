from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.excel_reader import read_excel_structure
from src.fill_plan_generator import (
    generate_fill_plan,
    write_config_error_report,
    write_fill_plan_json,
    write_fill_plan_xlsx,
)
from src.rule_parser import load_mapping_config
from src.template_analyzer import analyze_templates


def test_generate_fill_plan_contains_required_fields(tmp_path: Path) -> None:
    main_path = tmp_path / "主表.xlsx"
    template_path = tmp_path / "INVOICE.xlsx"
    config_path = tmp_path / "mapping_config.json"
    _make_main_workbook(main_path)
    _make_template_workbook(template_path)
    _write_valid_config(config_path)

    config, parse_errors = load_mapping_config(config_path)
    plan, plan_errors = generate_fill_plan(
        config=config,
        main_structure=read_excel_structure(main_path),
        template_structure=analyze_templates([template_path]),
        input_main_file=main_path,
        template_files=[template_path],
    )

    assert parse_errors == []
    assert plan_errors == []
    assert len(plan) == 6
    first = plan[0]
    assert first["source_file"] == "主表.xlsx"
    assert first["source_sheet"] == "发票数据"
    assert first["source_field"] == "客户名称"
    assert first["target_template"] == "INVOICE.xlsx"
    assert first["target_sheet"] == "Sheet1"
    assert first["target_cell"] == "B5"
    assert first["write_type"] == "single_value"
    assert first["is_calculated"] is False
    amount = [item for item in plan if item["source_field"] == "金额"][0]
    assert amount["is_calculated"] is True
    assert amount["calculation_formula"] == "数量 * 单价"


def test_fill_plan_writers_create_json_xlsx_and_error_report(tmp_path: Path) -> None:
    plan = [
        {
            "source_file": "主表.xlsx",
            "source_sheet": "发票数据",
            "source_field": "客户名称",
            "target_template": "INVOICE.xlsx",
            "output_name": "INVOICE_已填充.xlsx",
            "target_sheet": "Sheet1",
            "target_cell": "B5",
            "write_type": "single_value",
            "is_calculated": False,
            "calculation_formula": "",
            "required": True,
            "overwrite_formula": False,
            "risk_tips": [],
        }
    ]
    errors = [{"category": "target_cell", "location": "B0", "message": "invalid"}]

    json_path = tmp_path / "fill_plan.json"
    xlsx_path = tmp_path / "fill_plan.xlsx"
    error_path = tmp_path / "config_error_report.xlsx"
    write_fill_plan_json(plan, json_path)
    write_fill_plan_xlsx(plan, xlsx_path)
    write_config_error_report(errors, error_path)

    assert json.loads(json_path.read_text(encoding="utf-8"))["fill_plan"][0]["target_cell"] == "B5"
    assert load_workbook(xlsx_path).active["A2"].value == "主表.xlsx"
    assert load_workbook(error_path).active["A2"].value == "target_cell"


def _make_main_workbook(path: Path) -> None:
    workbook = Workbook()
    invoice = workbook.active
    invoice.title = "发票数据"
    invoice.append(["客户名称", "发票号"])
    invoice.append(["客户A", "INV-001"])
    detail = workbook.create_sheet("发票明细")
    detail.append(["品名", "数量", "单价"])
    detail.append(["产品A", 2, 10])
    workbook.save(path)


def _make_template_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["B5"] = ""
    sheet["F3"] = ""
    sheet["A10"] = "品名"
    sheet["C10"] = "数量"
    sheet["D10"] = "单价"
    sheet["E10"] = "金额"
    workbook.save(path)


def _write_valid_config(path: Path) -> None:
    config = {
        "project_name": "Excel 多模板自动填充系统",
        "main_file": "主表.xlsx",
        "templates": [
            {
                "template_name": "INVOICE.xlsx",
                "output_name": "INVOICE_已填充.xlsx",
                "source_sheet": "发票数据",
                "target_sheet": "Sheet1",
                "mappings": [
                    {
                        "source_field": "客户名称",
                        "target_cell": "B5",
                        "write_type": "single_value",
                        "required": True,
                        "overwrite_formula": False,
                    },
                    {
                        "source_field": "发票号",
                        "target_cell": "F3",
                        "write_type": "single_value",
                        "required": True,
                        "overwrite_formula": False,
                    },
                ],
                "detail_area": {
                    "source_sheet": "发票明细",
                    "target_sheet": "Sheet1",
                    "target_start_row": 10,
                    "target_end_row": 30,
                    "columns": {"品名": "A", "数量": "C", "单价": "D", "金额": "E"},
                    "calculations": [{"target_field": "金额", "formula": "数量 * 单价"}],
                },
            }
        ],
    }
    path.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
