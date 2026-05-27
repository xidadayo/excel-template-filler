from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.data_calculator import apply_calculations, calculate_expression, write_calculation_log, write_data_error_report
from src.data_cleaner import clean_value
from src.data_extractor import extract_data_from_plan


def test_clean_value_preserves_original_and_normalizes_values() -> None:
    cleaned = clean_value("  1,234.50 \n")
    assert cleaned["original_value"] == "  1,234.50 \n"
    assert cleaned["cleaned_value"] == 1234.5
    assert cleaned["is_empty"] is False

    empty = clean_value("  \n ")
    assert empty["cleaned_value"] is None
    assert empty["is_empty"] is True


def test_extracts_single_values_detail_rows_and_calculates(tmp_path: Path) -> None:
    main_path = tmp_path / "main.xlsx"
    _make_main(main_path)
    plan = [
        _single("发票数据", "客户名称", "B5"),
        _detail("发票明细", "品名", "A10", "A10:A30"),
        _detail("发票明细", "数量", "C10", "C10:C30"),
        _detail("发票明细", "单价", "D10", "D10:D30"),
        _detail("发票明细", "金额", "E10", "E10:E30", is_calculated=True, formula="数量 * 单价"),
    ]

    extracted, extract_errors = extract_data_from_plan(main_path, plan)
    calculated, logs, calculation_errors = apply_calculations(extracted)

    assert extract_errors == []
    assert calculation_errors == []
    assert calculated["items"][0]["cleaned_value"] == "客户A"
    assert calculated["items"][1]["rows"][0]["cleaned_value"] == "产品A"
    amount_item = calculated["items"][4]
    assert amount_item["rows"][0]["cleaned_value"] == 20
    assert amount_item["rows"][1]["cleaned_value"] == 37.5
    assert len(logs) == 2


def test_missing_field_is_reported_without_interrupting(tmp_path: Path) -> None:
    main_path = tmp_path / "main.xlsx"
    _make_main(main_path)
    plan = [_single("发票数据", "不存在字段", "B5")]

    extracted, errors = extract_data_from_plan(main_path, plan)

    assert extracted["items"][0]["status"] == "source_field_missing"
    assert errors[0]["category"] == "source_field"


def test_calculation_helpers_and_reports(tmp_path: Path) -> None:
    assert calculate_expression("数量 * 单价", {"数量": 3, "单价": 7}) == 21
    log_path = tmp_path / "calculation_log.xlsx"
    error_path = tmp_path / "data_error_report.xlsx"
    write_calculation_log(
        [
            {
                "plan_index": 1,
                "source_sheet": "发票明细",
                "target_field": "金额",
                "source_row": 2,
                "formula": "数量 * 单价",
                "input_values": "{'数量': 3, '单价': 7}",
                "result": 21,
                "status": "success",
                "message": "",
            }
        ],
        log_path,
    )
    write_data_error_report([{"category": "source_field", "location": "X", "message": "missing"}], error_path)

    assert load_workbook(log_path).active["G2"].value == 21
    assert load_workbook(error_path).active["A2"].value == "source_field"


def _make_main(path: Path) -> None:
    workbook = Workbook()
    invoice = workbook.active
    invoice.title = "发票数据"
    invoice.append(["客户名称", "发票号"])
    invoice.append([" 客户A ", "INV-001"])
    detail = workbook.create_sheet("发票明细")
    detail.append(["品名", "数量", "单价"])
    detail.append([" 产品A\n", "2", "10"])
    detail.append(["产品B", 3, "12.5"])
    workbook.save(path)


def _single(source_sheet: str, source_field: str, target_cell: str) -> dict[str, object]:
    return {
        "source_file": "main.xlsx",
        "source_sheet": source_sheet,
        "source_field": source_field,
        "target_template": "template.xlsx",
        "target_sheet": "Sheet1",
        "target_cell": target_cell,
        "write_type": "single_value",
        "required": True,
        "is_calculated": False,
    }


def _detail(
    source_sheet: str,
    source_field: str,
    target_cell: str,
    target_range: str,
    is_calculated: bool = False,
    formula: str = "",
) -> dict[str, object]:
    return {
        "source_file": "main.xlsx",
        "source_sheet": source_sheet,
        "source_field": source_field,
        "target_template": "template.xlsx",
        "target_sheet": "Sheet1",
        "target_cell": target_cell,
        "target_range": target_range,
        "write_type": "detail_column",
        "required": True,
        "is_calculated": is_calculated,
        "calculation_formula": formula,
    }
