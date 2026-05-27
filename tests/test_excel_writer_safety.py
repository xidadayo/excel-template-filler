from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from src.excel_writer import write_filled_templates


def test_writer_copies_template_preserves_style_and_writes_only_targets(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output_dir = tmp_path / "filled"
    _make_template(template)
    original_mtime = template.stat().st_mtime_ns
    extracted = {
        "items": [
            {
                "write_type": "single_value",
                "target_template": "template.xlsx",
                "output_name": "filled.xlsx",
                "target_sheet": "Sheet1",
                "target_cell": "B5",
                "cleaned_value": "Acme",
                "overwrite_formula": False,
            }
        ]
    }

    output_files, errors = write_filled_templates(extracted, [template], output_dir)

    assert errors == []
    assert len(output_files) == 1
    assert template.stat().st_mtime_ns == original_mtime

    original = load_workbook(template)
    filled = load_workbook(output_dir / "filled.xlsx")
    assert original["Sheet1"]["B5"].value is None
    assert filled["Sheet1"]["B5"].value == "Acme"
    assert filled["Sheet1"]["A1"].value == "Template"
    assert filled["Sheet1"]["A1"].font.bold is True
    assert filled["Sheet1"]["A1"].fill.fgColor.rgb == "00336699"
    assert filled["Sheet1"].column_dimensions["A"].width == 22
    assert filled["Sheet1"].row_dimensions[1].height == 28
    assert [str(item) for item in filled["Sheet1"].merged_cells.ranges] == ["D1:E1"]


def test_writer_does_not_overwrite_formula_by_default(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    _make_template(template)
    extracted = {
        "items": [
            {
                "write_type": "single_value",
                "target_template": "template.xlsx",
                "output_name": "filled.xlsx",
                "target_sheet": "Sheet1",
                "target_cell": "C5",
                "cleaned_value": 999,
                "overwrite_formula": False,
            }
        ]
    }

    _, errors = write_filled_templates(extracted, [template], tmp_path / "filled")

    filled = load_workbook(tmp_path / "filled" / "filled.xlsx", data_only=False)
    assert filled["Sheet1"]["C5"].value == "=SUM(A10:A30)"
    assert errors[0]["category"] == "formula"


def test_writer_records_detail_overflow_and_does_not_insert_rows(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    _make_template(template)
    extracted = {
        "items": [
            {
                "write_type": "detail_column",
                "target_template": "template.xlsx",
                "output_name": "filled.xlsx",
                "target_sheet": "Sheet1",
                "target_cell": "A10",
                "target_range": "A10:A11",
                "rows": [
                    {"cleaned_value": "row1"},
                    {"cleaned_value": "row2"},
                    {"cleaned_value": "row3"},
                ],
                "overwrite_formula": False,
            }
        ]
    }

    _, errors = write_filled_templates(extracted, [template], tmp_path / "filled")

    filled = load_workbook(tmp_path / "filled" / "filled.xlsx")
    assert filled["Sheet1"].max_row == 30
    assert filled["Sheet1"]["A10"].value == "row1"
    assert filled["Sheet1"]["A11"].value == "row2"
    assert filled["Sheet1"]["A12"].value is None
    assert errors[0]["category"] == "row_capacity"


def test_writer_supports_converted_template_alias(tmp_path: Path) -> None:
    converted_template = tmp_path / "legacy.xlsx"
    _make_template(converted_template)
    extracted = {
        "items": [
            {
                "write_type": "single_value",
                "target_template": "legacy.xls",
                "output_name": "filled.xlsx",
                "target_sheet": "Sheet1",
                "target_cell": "B5",
                "cleaned_value": "Alias OK",
                "overwrite_formula": False,
            }
        ]
    }

    outputs, errors = write_filled_templates(
        extracted,
        [converted_template],
        tmp_path / "filled",
        template_aliases={"legacy.xls": converted_template},
    )

    assert errors == []
    assert outputs[0].name == "filled.xlsx"
    assert load_workbook(outputs[0])["Sheet1"]["B5"].value == "Alias OK"


def _make_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Template"
    font = copy(sheet["A1"].font)
    font.bold = True
    font.size = 18
    sheet["A1"].font = font
    sheet["A1"].fill = PatternFill("solid", fgColor="336699")
    sheet["C5"] = "=SUM(A10:A30)"
    sheet["Z9"] = "do not touch"
    sheet["Z30"] = "row sentinel"
    sheet.merge_cells("D1:E1")
    sheet.column_dimensions["A"].width = 22
    sheet.row_dimensions[1].height = 28
    sheet["A30"] = None
    workbook.save(path)
