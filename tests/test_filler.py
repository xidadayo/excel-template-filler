from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_template_filler.config import load_config
from excel_template_filler.excel_ops import run_fill
from excel_template_filler.reports import RunReports


def test_fill_copies_template_and_preserves_style(tmp_path: Path) -> None:
    master_path = tmp_path / "master.xlsx"
    template_path = tmp_path / "template.xlsx"
    config_path = tmp_path / "config.json"

    _make_master(master_path, item_count=2)
    _make_template(template_path)
    _write_config(config_path, master_path, template_path, tmp_path / "outputs", max_rows=3)

    config = load_config(config_path)
    reports = RunReports(config.output_dir)
    run_fill(config, reports)
    reports.write()

    output_path = reports.files_dir / "filled.xlsx"
    assert output_path.exists()

    original = load_workbook(template_path)
    filled = load_workbook(output_path, data_only=False)
    assert original["Invoice"]["A1"].font.bold == filled["Invoice"]["A1"].font.bold
    assert original["Invoice"].column_dimensions["A"].width == filled["Invoice"].column_dimensions["A"].width
    assert filled["Invoice"]["B2"].value == "Acme"
    assert filled["Invoice"]["A5"].value == "S1"
    assert filled["Invoice"]["D20"].value == "=SUM(D5:D14)"
    assert reports.issues == []


def test_formula_target_is_not_overwritten_without_permission(tmp_path: Path) -> None:
    master_path = tmp_path / "master.xlsx"
    template_path = tmp_path / "template.xlsx"
    config_path = tmp_path / "config.json"

    _make_master(master_path, item_count=1)
    _make_template(template_path)
    config = {
        "master_path": str(master_path),
        "output_dir": str(tmp_path / "outputs"),
        "jobs": [
            {
                "name": "formula_guard",
                "template_path": str(template_path),
                "output_name": "filled.xlsx",
                "writes": [
                    {
                        "type": "cell",
                        "name": "try_formula",
                        "source": {"kind": "cell", "sheet": "Summary", "cell": "B2"},
                        "target": {"sheet": "Invoice", "cell": "D20"},
                    }
                ],
            }
        ],
    }
    config_path.write_text(json.dumps(config), encoding="utf-8")

    app_config = load_config(config_path)
    reports = RunReports(app_config.output_dir)
    run_fill(app_config, reports)

    output = load_workbook(reports.files_dir / "filled.xlsx", data_only=False)
    assert output["Invoice"]["D20"].value == "=SUM(D5:D14)"
    assert len(reports.issues) == 1
    assert "formula" in reports.issues[0].message


def test_range_overflow_is_reported_and_not_written(tmp_path: Path) -> None:
    master_path = tmp_path / "master.xlsx"
    template_path = tmp_path / "template.xlsx"
    config_path = tmp_path / "config.json"

    _make_master(master_path, item_count=4)
    _make_template(template_path)
    _write_config(config_path, master_path, template_path, tmp_path / "outputs", max_rows=2)

    config = load_config(config_path)
    reports = RunReports(config.output_dir)
    run_fill(config, reports)

    output = load_workbook(reports.files_dir / "filled.xlsx")
    assert output["Invoice"]["A5"].value is None
    assert len(reports.issues) == 1
    assert "auto insert is disabled" in reports.issues[0].message


def _make_master(path: Path, item_count: int) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Field", "Value"])
    summary.append(["customer", "Acme"])
    items = wb.create_sheet("Items")
    items.append(["SKU", "Name", "Qty", "Price"])
    for index in range(1, item_count + 1):
        items.append([f"S{index}", f"Item {index}", index, index * 10])
    wb.save(path)


def _make_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws["A1"] = "Invoice"
    font = copy(ws["A1"].font)
    font.bold = True
    ws["A1"].font = font
    ws["B2"] = ""
    ws["D20"] = "=SUM(D5:D14)"
    ws.column_dimensions["A"].width = 22
    wb.save(path)


def _write_config(config_path: Path, master_path: Path, template_path: Path, output_dir: Path, max_rows: int) -> None:
    config = {
        "master_path": str(master_path),
        "output_dir": str(output_dir),
        "jobs": [
            {
                "name": "invoice",
                "template_path": str(template_path),
                "output_name": "filled.xlsx",
                "writes": [
                    {
                        "type": "cell",
                        "name": "customer",
                        "source": {
                            "kind": "lookup",
                            "sheet": "Summary",
                            "key_column": "Field",
                            "key_value": "customer",
                            "value_column": "Value",
                        },
                        "target": {"sheet": "Invoice", "cell": "B2"},
                    },
                    {
                        "type": "range",
                        "name": "items",
                        "source": {
                            "kind": "table",
                            "sheet": "Items",
                            "columns": ["SKU", "Name", "Qty", "Price"],
                        },
                        "target": {"sheet": "Invoice", "start_cell": "A5", "max_rows": max_rows},
                    },
                ],
            }
        ],
    }
    config_path.write_text(json.dumps(config), encoding="utf-8")
