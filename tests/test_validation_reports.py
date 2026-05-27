from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.report_writer import write_error_report, write_process_log, write_summary, write_validation_report
from src.validator import validate_outputs


def test_validator_detects_output_integrity_and_changed_non_targets(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output = tmp_path / "filled.xlsx"
    _make_template(template)
    _make_template(output)
    filled = load_workbook(output)
    ws = filled["Sheet1"]
    ws["B5"] = "Acme"
    ws["A10"] = "Item 1"
    ws["A11"] = "Item 2"
    ws["Z9"] = "changed"
    filled.save(output)

    extracted = {
        "items": [
            {
                "write_type": "single_value",
                "target_template": "template.xlsx",
                "output_name": "filled.xlsx",
                "target_sheet": "Sheet1",
                "target_cell": "B5",
                "source_field": "客户名称",
                "cleaned_value": "Acme",
                "required": True,
            },
            {
                "write_type": "detail_column",
                "target_template": "template.xlsx",
                "output_name": "filled.xlsx",
                "target_sheet": "Sheet1",
                "target_cell": "A10",
                "target_range": "A10:A11",
                "source_field": "品名",
                "rows": [{"cleaned_value": "Item 1"}, {"cleaned_value": "Item 2"}],
                "required": True,
            },
        ]
    }

    rows, errors = validate_outputs(extracted, [template], [output])

    assert any(row["check"] == "output_open" and row["status"] == "passed" for row in rows)
    assert any(error["category"] == "non_target_cells" for error in errors)


def test_report_writer_outputs_workbooks_and_summary(tmp_path: Path) -> None:
    validation_path = tmp_path / "validation_report.xlsx"
    error_path = tmp_path / "error_report.xlsx"
    process_path = tmp_path / "process_log.xlsx"
    summary_path = tmp_path / "summary.txt"

    write_validation_report([{"check": "output_open", "location": "filled.xlsx", "status": "passed", "message": "ok"}], validation_path)
    write_error_report([{"category": "x", "location": "y", "message": "z"}], error_path)
    write_process_log([{"step": "plan", "status": "done", "message": "ok"}], process_path)
    write_summary(
        summary_path,
        main_file="main.xlsx",
        templates=["template.xlsx"],
        extracted_data={
            "items": [
                {
                    "target_template": "template.xlsx",
                    "source_field": "客户名称",
                    "is_empty": False,
                }
            ]
        },
        output_files=[tmp_path / "filled.xlsx"],
        errors=[{"category": "x", "location": "y", "message": "z"}],
        validation_rows=[{"check": "output_open", "location": "filled.xlsx", "status": "passed", "message": "ok"}],
    )

    assert load_workbook(validation_path).active["A2"].value == "output_open"
    assert load_workbook(error_path).active["A2"].value == "x"
    assert load_workbook(process_path).active["A2"].value == "plan"
    summary = summary_path.read_text(encoding="utf-8")
    assert "main_file: main.xlsx" in summary
    assert "template.xlsx: 客户名称" in summary


def _make_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Template"
    sheet["C5"] = "=SUM(A10:A11)"
    sheet["Z9"] = "sentinel"
    sheet.column_dimensions["A"].width = 20
    workbook.save(path)
