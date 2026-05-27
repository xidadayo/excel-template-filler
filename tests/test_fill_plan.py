from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

from src.file_manager import FileManager
from src.xls_converter import ConversionRecord


def test_file_manager_collects_one_main_and_templates(tmp_path: Path) -> None:
    main_dir = tmp_path / "input" / "main_excel"
    template_dir = tmp_path / "input" / "templates"
    main_dir.mkdir(parents=True)
    template_dir.mkdir(parents=True)
    _make_workbook(main_dir / "main.xlsx")
    _make_workbook(template_dir / "template.xlsx")

    inputs = FileManager(main_dir, template_dir, tmp_path / "output").collect_inputs()

    assert inputs.main_excel.name == "main.xlsx"
    assert [path.name for path in inputs.templates] == ["template.xlsx"]


def test_file_manager_records_failed_xls_template_conversion(tmp_path: Path, monkeypatch) -> None:
    main_dir = tmp_path / "input" / "main_excel"
    template_dir = tmp_path / "input" / "templates"
    main_dir.mkdir(parents=True)
    template_dir.mkdir(parents=True)
    _make_workbook(main_dir / "main.xlsx")
    _make_workbook(template_dir / "template.xlsx")
    (template_dir / "legacy.xls").write_bytes(b"not a real xls")

    def fake_convert(source_path: Path, output_dir: Path, backup_dir: Path) -> ConversionRecord:
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / source_path.name
        backup_path.write_bytes(source_path.read_bytes())
        return ConversionRecord(str(source_path), str(backup_path), str(output_dir / "legacy.xlsx"), "fake", "failed", "conversion failed")

    monkeypatch.setattr("src.file_manager.convert_xls_file", fake_convert)

    inputs = FileManager(main_dir, template_dir, tmp_path / "output").collect_inputs()

    assert inputs.main_excel.name == "main.xlsx"
    assert [path.name for path in inputs.templates] == ["template.xlsx"]
    assert inputs.conversion_records[0].status == "failed"
    assert (tmp_path / "output" / "backups" / "legacy.xls").exists()
    assert (tmp_path / "output" / "logs" / "conversion_log.xlsx").exists()


def test_file_manager_reports_failed_main_xls_conversion(tmp_path: Path, monkeypatch) -> None:
    main_dir = tmp_path / "input" / "main_excel"
    template_dir = tmp_path / "input" / "templates"
    main_dir.mkdir(parents=True)
    template_dir.mkdir(parents=True)
    (main_dir / "legacy.xls").write_bytes(b"not a real xls")
    _make_workbook(template_dir / "template.xlsx")

    def fake_convert(source_path: Path, output_dir: Path, backup_dir: Path) -> ConversionRecord:
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / source_path.name
        backup_path.write_bytes(source_path.read_bytes())
        return ConversionRecord(str(source_path), str(backup_path), str(output_dir / "legacy.xlsx"), "fake", "failed", "conversion failed")

    monkeypatch.setattr("src.file_manager.convert_xls_file", fake_convert)

    try:
        FileManager(main_dir, template_dir, tmp_path / "output").collect_inputs()
    except ValueError as exc:
        assert "main .xls conversion failed" in str(exc)
    else:
        raise AssertionError("Expected ValueError for failed main .xls conversion")


def test_file_manager_uses_converted_xls_template_alias(tmp_path: Path, monkeypatch) -> None:
    main_dir = tmp_path / "input" / "main_excel"
    template_dir = tmp_path / "input" / "templates"
    main_dir.mkdir(parents=True)
    template_dir.mkdir(parents=True)
    _make_workbook(main_dir / "main.xlsx")
    (template_dir / "legacy.xls").write_bytes(b"fake xls")

    def fake_convert(source_path: Path, output_dir: Path, backup_dir: Path) -> ConversionRecord:
        output_dir.mkdir(parents=True, exist_ok=True)
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / source_path.name
        backup_path.write_bytes(source_path.read_bytes())
        converted = output_dir / "legacy.xlsx"
        _make_workbook(converted)
        return ConversionRecord(str(source_path), str(backup_path), str(converted), "fake", "converted", "ok")

    monkeypatch.setattr("src.file_manager.convert_xls_file", fake_convert)

    inputs = FileManager(main_dir, template_dir, tmp_path / "output").collect_inputs()

    assert [path.name for path in inputs.templates] == ["legacy.xlsx"]
    assert inputs.template_aliases["legacy.xls"].name == "legacy.xlsx"
    assert load_workbook(tmp_path / "output" / "logs" / "conversion_log.xlsx").active["E2"].value == "converted"


def _make_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "ok"
    workbook.save(path)
