from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]


def main() -> None:
    data_dir = ROOT / "examples" / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    master_path = data_dir / "master.xlsx"
    template_path = data_dir / "invoice_template.xlsx"
    config_path = ROOT / "configs" / "example_config.json"

    master = Workbook()
    ws = master.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    ws.append(["customer_name", "Acme Trading"])
    ws.append(["order_no", "SO-2026-001"])
    items = master.create_sheet("Items")
    items.append(["SKU", "Name", "Qty", "Price"])
    items.append(["A001", "Canvas Tote", 3, 19.9])
    items.append(["B002", "Travel Pouch", 5, 8.5])
    master.save(master_path)

    template = Workbook()
    invoice = template.active
    invoice.title = "Invoice"
    invoice["A1"] = "Invoice"
    invoice["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    invoice["A1"].fill = PatternFill("solid", fgColor="336699")
    invoice["A3"] = "Customer"
    invoice["A4"] = "Order No"
    invoice["A8"] = "SKU"
    invoice["B8"] = "Name"
    invoice["C8"] = "Qty"
    invoice["D8"] = "Price"
    invoice["D20"] = "=SUM(D9:D18)"
    invoice.column_dimensions["A"].width = 18
    invoice.column_dimensions["B"].width = 24
    template.save(template_path)

    config = {
        "master_path": "../examples/data/master.xlsx",
        "output_dir": "../outputs",
        "jobs": [
            {
                "name": "invoice",
                "template_path": "../examples/data/invoice_template.xlsx",
                "output_name": "invoice_filled.xlsx",
                "writes": [
                    {
                        "type": "cell",
                        "name": "fill_customer",
                        "source": {
                            "kind": "lookup",
                            "sheet": "Summary",
                            "key_column": "Field",
                            "key_value": "customer_name",
                            "value_column": "Value"
                        },
                        "target": {"sheet": "Invoice", "cell": "B3"}
                    },
                    {
                        "type": "cell",
                        "name": "fill_order_no",
                        "source": {
                            "kind": "lookup",
                            "sheet": "Summary",
                            "key_column": "Field",
                            "key_value": "order_no",
                            "value_column": "Value"
                        },
                        "target": {"sheet": "Invoice", "cell": "B4"}
                    },
                    {
                        "type": "range",
                        "name": "fill_items",
                        "source": {
                            "kind": "table",
                            "sheet": "Items",
                            "columns": ["SKU", "Name", "Qty", "Price"],
                            "header_row": 1,
                            "start_row": 2
                        },
                        "target": {"sheet": "Invoice", "start_cell": "A9", "max_rows": 10}
                    }
                ]
            }
        ]
    }
    config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Created {master_path}")
    print(f"Created {template_path}")
    print(f"Created {config_path}")


if __name__ == "__main__":
    main()
