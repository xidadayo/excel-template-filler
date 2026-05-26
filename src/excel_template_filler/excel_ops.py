from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from .models import AppConfig, CellSource, CellWriteRule, LookupSource, RangeWriteRule, TableSource
from .reports import Issue, RunReports, ValidationItem


def run_fill(config: AppConfig, reports: RunReports) -> None:
    master_wb = load_workbook(config.master_path, data_only=False)
    logger.info("Loaded master workbook: {}", config.master_path)

    for job in config.jobs:
        if not job.template_path.exists():
            reports.add_issue(Issue("error", job.name, "__template__", f"template_path does not exist: {job.template_path}"))
            continue

        output_path = reports.files_dir / job.output_name
        shutil.copy2(job.template_path, output_path)
        logger.info("Copied template '{}' to '{}'", job.template_path, output_path)

        template_wb = load_workbook(output_path)
        for rule in job.writes:
            try:
                if isinstance(rule, CellWriteRule):
                    _apply_cell_rule(master_wb, template_wb, job.name, rule, reports)
                elif isinstance(rule, RangeWriteRule):
                    _apply_range_rule(master_wb, template_wb, job.name, rule, reports)
            except Exception as exc:
                logger.exception("Rule failed: {} / {}", job.name, rule.name)
                reports.add_issue(Issue("error", job.name, rule.name, str(exc)))

        template_wb.save(output_path)
        reports.add_output(output_path)
        logger.info("Saved output workbook: {}", output_path)


def _apply_cell_rule(master_wb: Any, template_wb: Any, job_name: str, rule: CellWriteRule, reports: RunReports) -> None:
    value = _read_cell_source(master_wb, rule.source, job_name, rule.name, reports)
    if value is _MISSING:
        return

    ws = _get_sheet(template_wb, rule.target.sheet, job_name, rule.name, reports)
    if ws is None:
        return

    target = ws[rule.target.cell]
    if not _can_write_cell(target, rule.allow_overwrite_formula, job_name, rule.name, reports):
        return

    target.value = value
    reports.add_validation(ValidationItem(job_name, rule.name, "written", "cell value written", f"{rule.target.sheet}!{rule.target.cell}"))


def _apply_range_rule(master_wb: Any, template_wb: Any, job_name: str, rule: RangeWriteRule, reports: RunReports) -> None:
    rows = _read_table_source(master_wb, rule.source, job_name, rule.name, reports)
    if rows is _MISSING:
        return

    if len(rows) > rule.target.max_rows:
        reports.add_issue(
            Issue(
                "error",
                job_name,
                rule.name,
                f"target range has {rule.target.max_rows} rows but source has {len(rows)} rows; auto insert is disabled",
                f"{rule.target.sheet}!{rule.target.start_cell}",
            )
        )
        return

    ws = _get_sheet(template_wb, rule.target.sheet, job_name, rule.name, reports)
    if ws is None:
        return

    start_col_letter, start_row = coordinate_from_string(rule.target.start_cell)
    start_col = column_index_from_string(start_col_letter)
    for row_offset, row in enumerate(rows):
        for col_offset, column_name in enumerate(rule.source.columns):
            cell = ws.cell(row=start_row + row_offset, column=start_col + col_offset)
            if not _can_write_cell(cell, rule.allow_overwrite_formula, job_name, rule.name, reports):
                return
            cell.value = row[column_name]

    end_col = get_column_letter(start_col + len(rule.source.columns) - 1)
    end_row = start_row + max(len(rows), 1) - 1
    reports.add_validation(
        ValidationItem(job_name, rule.name, "written", f"{len(rows)} rows written", f"{rule.target.sheet}!{rule.target.start_cell}:{end_col}{end_row}")
    )


def _read_cell_source(master_wb: Any, source: LookupSource | CellSource, job_name: str, rule_name: str, reports: RunReports) -> Any:
    ws = _get_sheet(master_wb, source.sheet, job_name, rule_name, reports)
    if ws is None:
        return _MISSING

    if isinstance(source, CellSource):
        return ws[source.cell].value

    headers = _headers_by_name(ws, source.header_row)
    missing = [name for name in (source.key_column, source.value_column) if name not in headers]
    if missing:
        reports.add_issue(Issue("error", job_name, rule_name, f"missing source column(s): {', '.join(missing)}", source.sheet))
        return _MISSING

    key_idx = headers[source.key_column]
    value_idx = headers[source.value_column]
    for row in ws.iter_rows(min_row=source.header_row + 1):
        if row[key_idx - 1].value == source.key_value:
            return row[value_idx - 1].value

    reports.add_issue(Issue("error", job_name, rule_name, f"lookup key not found: {source.key_value}", source.sheet))
    return _MISSING


def _read_table_source(master_wb: Any, source: TableSource, job_name: str, rule_name: str, reports: RunReports) -> list[dict[str, Any]] | object:
    ws = _get_sheet(master_wb, source.sheet, job_name, rule_name, reports)
    if ws is None:
        return _MISSING

    headers = _headers_by_name(ws, source.header_row)
    missing = [name for name in source.columns if name not in headers]
    if missing:
        reports.add_issue(Issue("error", job_name, rule_name, f"missing source column(s): {', '.join(missing)}", source.sheet))
        return _MISSING

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=source.start_row):
        values = {column_name: row[headers[column_name] - 1].value for column_name in source.columns}
        if all(value is None for value in values.values()):
            continue
        rows.append(values)
    return rows


def _headers_by_name(ws: Any, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            headers[str(cell.value)] = cell.column
    return headers


def _get_sheet(workbook: Any, sheet_name: str, job_name: str, rule_name: str, reports: RunReports) -> Any | None:
    if sheet_name not in workbook.sheetnames:
        reports.add_issue(Issue("error", job_name, rule_name, f"sheet not found: {sheet_name}", sheet_name))
        return None
    return workbook[sheet_name]


def _can_write_cell(cell: Any, allow_overwrite_formula: bool, job_name: str, rule_name: str, reports: RunReports) -> bool:
    location = f"{cell.parent.title}!{cell.coordinate}"
    if not isinstance(cell, Cell):
        reports.add_issue(Issue("error", job_name, rule_name, "target cell is not writable, possibly a non-anchor merged cell", location))
        return False
    if cell.data_type == "f" and not allow_overwrite_formula:
        reports.add_issue(Issue("error", job_name, rule_name, "target cell contains formula and overwrite is not allowed", location))
        return False
    return True


class _Missing:
    pass


_MISSING = _Missing()
