"""Validate filled workbook outputs against templates and extracted data."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


def validate_outputs(
    extracted_data: dict[str, Any],
    template_files: list[Path],
    output_files: list[Path],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    template_by_name = {path.name: path for path in template_files}
    output_by_name = {path.name: path for path in output_files}
    items_by_template = _group_items(extracted_data.get("items", []))
    validation_rows: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []

    for template_name, items in items_by_template.items():
        output_name = _output_name(items, template_name)
        template_path = template_by_name.get(template_name)
        output_path = output_by_name.get(output_name)

        if template_path is None:
            _add(validation_rows, errors, "template_exists", template_name, "failed", "original template not found")
            continue
        if output_path is None:
            _add(validation_rows, errors, "output_exists", output_name, "failed", "filled output file not found")
            continue

        try:
            template_wb = load_workbook(template_path, data_only=False)
            output_wb = load_workbook(output_path, data_only=False)
        except Exception as exc:
            _add(validation_rows, errors, "output_open", output_name, "failed", f"workbook cannot be opened: {exc}")
            continue

        try:
            _add(validation_rows, errors, "original_template_open", template_name, "passed", "original template opens and is used as the unchanged baseline", report_error=False)
            _validate_workbook(template_wb, output_wb, items, template_name, output_name, validation_rows, errors)
        finally:
            template_wb.close()
            output_wb.close()

    return validation_rows, errors


def _validate_workbook(
    template_wb: Any,
    output_wb: Any,
    items: list[dict[str, Any]],
    template_name: str,
    output_name: str,
    rows: list[dict[str, Any]],
    errors: list[dict[str, str]],
) -> None:
    _add(rows, errors, "output_open", output_name, "passed", "filled output file opens successfully", report_error=False)
    if template_wb.sheetnames != output_wb.sheetnames:
        _add(rows, errors, "sheet_structure", output_name, "failed", "sheet names/order changed")
    else:
        _add(rows, errors, "sheet_structure", output_name, "passed", "sheet names/order unchanged", report_error=False)

    targets = _target_cells(items)
    formula_overwrite_allowed = _formula_overwrite_cells(items)
    _validate_required_and_counts(output_wb, items, rows, errors)
    _validate_template_structure(template_wb, output_wb, output_name, rows, errors)
    _validate_non_target_cells(template_wb, output_wb, targets, output_name, rows, errors)
    _validate_formula_cells(template_wb, output_wb, formula_overwrite_allowed, output_name, rows, errors)
    _validate_totals(items, template_name, rows, errors)


def _validate_required_and_counts(output_wb: Any, items: list[dict[str, Any]], rows: list[dict[str, Any]], errors: list[dict[str, str]]) -> None:
    for item in items:
        if item.get("write_type") == "single_value":
            location = _item_location(item)
            if item.get("required") and item.get("cleaned_value") in (None, ""):
                _add(rows, errors, "required_field", location, "failed", "required field is empty")
                continue
            cell = output_wb[str(item.get("target_sheet"))][str(item.get("target_cell"))]
            if item.get("required") and cell.value in (None, ""):
                _add(rows, errors, "required_field", location, "failed", "target cell is empty after fill")
            else:
                _add(rows, errors, "required_field", location, "passed", "required field written", report_error=False)
        elif item.get("write_type") == "detail_column":
            item_rows = item.get("rows") or []
            target_range = str(item.get("target_range") or "")
            capacity = _range_capacity(target_range)
            written_count = min(len(item_rows), capacity)
            if len(item_rows) > capacity:
                _add(rows, errors, "detail_row_count", target_range, "failed", f"source rows={len(item_rows)}, writable rows={capacity}")
            else:
                _add(rows, errors, "detail_row_count", target_range, "passed", f"written rows={written_count}", report_error=False)


def _validate_template_structure(template_wb: Any, output_wb: Any, output_name: str, rows: list[dict[str, Any]], errors: list[dict[str, str]]) -> None:
    for sheet_name in template_wb.sheetnames:
        if sheet_name not in output_wb.sheetnames:
            continue
        template_ws = template_wb[sheet_name]
        output_ws = output_wb[sheet_name]
        if [str(item) for item in template_ws.merged_cells.ranges] != [str(item) for item in output_ws.merged_cells.ranges]:
            _add(rows, errors, "merged_cells", f"{output_name}:{sheet_name}", "failed", "merged cell ranges changed")
        else:
            _add(rows, errors, "merged_cells", f"{output_name}:{sheet_name}", "passed", "merged cell ranges unchanged", report_error=False)
        if _dimensions_changed(template_ws, output_ws):
            _add(rows, errors, "dimensions", f"{output_name}:{sheet_name}", "failed", "row heights or column widths changed")
        else:
            _add(rows, errors, "dimensions", f"{output_name}:{sheet_name}", "passed", "row heights and column widths unchanged", report_error=False)


def _validate_non_target_cells(
    template_wb: Any,
    output_wb: Any,
    targets: set[tuple[str, str]],
    output_name: str,
    rows: list[dict[str, Any]],
    errors: list[dict[str, str]],
) -> None:
    changed = []
    for sheet_name in template_wb.sheetnames:
        if sheet_name not in output_wb.sheetnames:
            continue
        template_ws = template_wb[sheet_name]
        output_ws = output_wb[sheet_name]
        max_row = max(template_ws.max_row, output_ws.max_row)
        max_col = max(template_ws.max_column, output_ws.max_column)
        for row in range(1, max_row + 1):
            for column in range(1, max_col + 1):
                coordinate = output_ws.cell(row=row, column=column).coordinate
                if (sheet_name, coordinate) in targets:
                    continue
                if template_ws.cell(row=row, column=column).value != output_ws.cell(row=row, column=column).value:
                    changed.append(f"{sheet_name}!{coordinate}")
    if changed:
        _add(rows, errors, "non_target_cells", output_name, "failed", "non-target cells changed: " + ", ".join(changed[:20]))
    else:
        _add(rows, errors, "non_target_cells", output_name, "passed", "non-target cell values unchanged", report_error=False)


def _validate_formula_cells(
    template_wb: Any,
    output_wb: Any,
    allowed: set[tuple[str, str]],
    output_name: str,
    rows: list[dict[str, Any]],
    errors: list[dict[str, str]],
) -> None:
    overwritten = []
    for sheet_name in template_wb.sheetnames:
        if sheet_name not in output_wb.sheetnames:
            continue
        template_ws = template_wb[sheet_name]
        output_ws = output_wb[sheet_name]
        for row in template_ws.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                if (sheet_name, cell.coordinate) in allowed:
                    continue
                if output_ws[cell.coordinate].value != cell.value:
                    overwritten.append(f"{sheet_name}!{cell.coordinate}")
    if overwritten:
        _add(rows, errors, "formula_cells", output_name, "failed", "formula cells overwritten: " + ", ".join(overwritten))
    else:
        _add(rows, errors, "formula_cells", output_name, "passed", "formula cells unchanged unless explicitly allowed", report_error=False)


def _validate_totals(items: list[dict[str, Any]], template_name: str, rows: list[dict[str, Any]], errors: list[dict[str, str]]) -> None:
    for keyword, check_name in (("数量", "total_quantity"), ("金额", "total_amount")):
        relevant = [item for item in items if item.get("write_type") == "detail_column" and keyword in str(item.get("source_field"))]
        if not relevant:
            continue
        for item in relevant:
            source_values = [_number(row.get("cleaned_value")) for row in item.get("rows", []) if _number(row.get("cleaned_value")) is not None]
            target_range = str(item.get("target_range") or "")
            capacity = _range_capacity(target_range)
            written_values = source_values[:capacity]
            if sum(source_values) != sum(written_values):
                _add(rows, errors, check_name, f"{template_name}:{target_range}", "failed", f"source total={sum(source_values)}, written total={sum(written_values)}")
            else:
                _add(rows, errors, check_name, f"{template_name}:{target_range}", "passed", f"total={sum(source_values)}", report_error=False)


def _target_cells(items: list[dict[str, Any]]) -> set[tuple[str, str]]:
    targets: set[tuple[str, str]] = set()
    for item in items:
        sheet_name = str(item.get("target_sheet"))
        if item.get("write_type") == "single_value" and item.get("target_cell"):
            targets.add((sheet_name, str(item.get("target_cell"))))
        elif item.get("write_type") == "detail_column" and item.get("target_range"):
            min_col, min_row, max_col, max_row = range_boundaries(str(item.get("target_range")))
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    targets.add((sheet_name, f"{get_column_letter(column)}{row}"))
    return targets


def _formula_overwrite_cells(items: list[dict[str, Any]]) -> set[tuple[str, str]]:
    return {
        (str(item.get("target_sheet")), str(item.get("target_cell")))
        for item in items
        if item.get("overwrite_formula") and item.get("target_cell")
    }


def _group_items(items: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in items:
        template = item.get("target_template")
        if template:
            grouped.setdefault(str(template), []).append(item)
    return grouped


def _output_name(items: list[dict[str, Any]], template_name: str) -> str:
    for item in items:
        if item.get("output_name"):
            return str(item.get("output_name"))
    template_path = Path(template_name)
    return f"{template_path.stem}_filled{template_path.suffix}"


def _range_capacity(target_range: str) -> int:
    if not target_range:
        return 0
    _, min_row, _, max_row = range_boundaries(target_range)
    return max_row - min_row + 1


def _dimensions_changed(template_ws: Any, output_ws: Any) -> bool:
    template_cols = {key: dim.width for key, dim in template_ws.column_dimensions.items()}
    output_cols = {key: dim.width for key, dim in output_ws.column_dimensions.items()}
    template_rows = {key: dim.height for key, dim in template_ws.row_dimensions.items()}
    output_rows = {key: dim.height for key, dim in output_ws.row_dimensions.items()}
    return template_cols != output_cols or template_rows != output_rows


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def _item_location(item: dict[str, Any]) -> str:
    return f"{item.get('target_template')}:{item.get('target_sheet')}!{item.get('target_cell')}"


def _add(
    rows: list[dict[str, Any]],
    errors: list[dict[str, str]],
    check_name: str,
    location: str,
    status: str,
    message: str,
    report_error: bool = True,
) -> None:
    rows.append({"check": check_name, "location": location, "status": status, "message": message})
    if status != "passed" and report_error:
        errors.append({"category": check_name, "location": location, "message": message})
