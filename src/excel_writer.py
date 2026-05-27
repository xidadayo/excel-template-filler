"""Safe Excel template writer.

This module writes values only to copied templates and only mutates cell.value.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import range_boundaries

from src.exception_handler import error


def write_filled_templates(
    extracted_data: dict[str, Any],
    template_files: list[Path],
    output_dir: str | Path,
    allow_insert_rows: bool = False,
    template_aliases: dict[str, Path] | None = None,
) -> tuple[list[Path], list[dict[str, str]]]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    templates_by_name = {path.name: path for path in template_files}
    templates_by_name.update(template_aliases or {})
    items_by_template = _group_items_by_template(extracted_data.get("items", []))
    errors: list[dict[str, str]] = []
    output_files: list[Path] = []

    for template_name, items in items_by_template.items():
        template_path = templates_by_name.get(template_name)
        if template_path is None:
            errors.append(error("template", template_name, "template file not found; write skipped"))
            continue

        output_name = _output_name(items, template_path)
        copied_path = output_path / output_name
        shutil.copy2(template_path, copied_path)
        workbook = load_workbook(copied_path, keep_vba=copied_path.suffix.lower() == ".xlsm")

        try:
            for item in items:
                if item.get("status"):
                    errors.append(error("data", _item_location(item), f"item status is {item.get('status')}; write skipped"))
                    continue
                if item.get("write_type") == "single_value":
                    _write_single_value(workbook, item, errors)
                elif item.get("write_type") == "detail_column":
                    _write_detail_column(workbook, item, errors, allow_insert_rows)
            workbook.save(copied_path)
            output_files.append(copied_path)
        finally:
            workbook.close()

    return output_files, errors


def _group_items_by_template(items: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in items:
        template_name = item.get("target_template")
        if not template_name:
            continue
        grouped.setdefault(str(template_name), []).append(item)
    return grouped


def _output_name(items: list[dict[str, Any]], template_path: Path) -> str:
    for item in items:
        output_name = item.get("output_name")
        if output_name:
            return str(output_name)
    return f"{template_path.stem}_filled{template_path.suffix}"


def _write_single_value(workbook: Any, item: dict[str, Any], errors: list[dict[str, str]]) -> None:
    worksheet = _worksheet(workbook, item, errors)
    if worksheet is None:
        return
    cell_ref = str(item.get("target_cell") or "")
    cell = worksheet[cell_ref]
    if not _can_write_cell(cell, item, errors):
        return
    cell.value = item.get("cleaned_value")


def _write_detail_column(
    workbook: Any,
    item: dict[str, Any],
    errors: list[dict[str, str]],
    allow_insert_rows: bool,
) -> None:
    worksheet = _worksheet(workbook, item, errors)
    if worksheet is None:
        return

    target_range = item.get("target_range")
    if not target_range:
        errors.append(error("target_range", _item_location(item), "target_range is missing"))
        return
    min_col, min_row, max_col, max_row = range_boundaries(str(target_range))
    if min_col != max_col:
        errors.append(error("target_range", str(target_range), "detail target range must be a single column"))
        return

    rows = item.get("rows") or []
    capacity = max_row - min_row + 1
    if len(rows) > capacity and not allow_insert_rows:
        errors.append(
            error(
                "row_capacity",
                str(target_range),
                f"detail rows exceed target range: source rows={len(rows)}, capacity={capacity}; extra rows were not written",
            )
        )

    for offset, row in enumerate(rows[:capacity]):
        cell = worksheet.cell(row=min_row + offset, column=min_col)
        if not _can_write_cell(cell, item, errors):
            continue
        cell.value = row.get("cleaned_value")


def _worksheet(workbook: Any, item: dict[str, Any], errors: list[dict[str, str]]) -> Any | None:
    sheet_name = item.get("target_sheet")
    if sheet_name not in workbook.sheetnames:
        errors.append(error("target_sheet", str(sheet_name), "target sheet not found; write skipped"))
        return None
    return workbook[str(sheet_name)]


def _can_write_cell(cell: Any, item: dict[str, Any], errors: list[dict[str, str]]) -> bool:
    location = f"{cell.parent.title}!{cell.coordinate}"
    if not isinstance(cell, Cell):
        errors.append(error("merged_cell", location, "target is not a writable cell, possibly a non-anchor merged cell"))
        return False
    if isinstance(cell.value, str) and cell.value.startswith("=") and not item.get("overwrite_formula"):
        errors.append(error("formula", location, "target cell contains formula and overwrite_formula is false"))
        return False
    return True


def _item_location(item: dict[str, Any]) -> str:
    return f"{item.get('target_template')}:{item.get('target_sheet')}!{item.get('target_cell')}"
