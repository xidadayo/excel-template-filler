"""Read-only Excel workbook structure analysis."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_excel_structure(path: str | Path) -> dict[str, Any]:
    excel_path = Path(path)
    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            header_row, fields = detect_header_row(worksheet)
            sheets.append(
                {
                    "sheet_name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "header_row": header_row,
                    "fields": fields,
                }
            )
        return {
            "file_name": excel_path.name,
            "file_path": str(excel_path),
            "sheet_names": workbook.sheetnames,
            "sheets": sheets,
        }
    finally:
        workbook.close()


def detect_header_row(worksheet: Any, scan_rows: int = 20) -> tuple[int | None, list[str]]:
    best_row_index: int | None = None
    best_values: list[str] = []
    best_score = 0

    max_scan_row = min(worksheet.max_row or 0, scan_rows)
    for row in worksheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        string_like_count = sum(1 for value in values if not _looks_numeric(value))
        score = len(values) + string_like_count
        if len(values) >= 2 and score > best_score:
            best_score = score
            best_values = values
            best_row_index = 0

    if best_row_index is None:
        return _first_non_empty_row(worksheet, max_scan_row)

    # Read-only worksheets with values_only=True do not expose row numbers, so rescan.
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
        start=1,
    ):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if values == best_values:
            return row_index, best_values
    return None, best_values


def _first_non_empty_row(worksheet: Any, max_scan_row: int) -> tuple[int | None, list[str]]:
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
        start=1,
    ):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if values:
            return row_index, values
    return None, []


def _looks_numeric(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False
