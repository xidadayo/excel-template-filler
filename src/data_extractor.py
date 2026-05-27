"""Extract data from the main workbook according to a fill plan."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.data_cleaner import clean_row, clean_value
from src.excel_reader import detect_header_row


def extract_data_from_plan(main_excel: str | Path, fill_plan: list[dict[str, Any]]) -> tuple[dict[str, Any], list[dict[str, str]]]:
    workbook = load_workbook(main_excel, read_only=True, data_only=True)
    errors: list[dict[str, str]] = []
    sheet_cache: dict[str, dict[str, Any]] = {}
    items: list[dict[str, Any]] = []

    try:
        for index, plan_item in enumerate(fill_plan, start=1):
            source_sheet = plan_item.get("source_sheet", "")
            source_field = plan_item.get("source_field", "")
            write_type = plan_item.get("write_type", "")
            sheet_data = _get_sheet_data(workbook, source_sheet, sheet_cache, errors)
            if sheet_data is None:
                items.append(_empty_item(index, plan_item, "source_sheet_missing"))
                continue

            if source_field not in sheet_data["fields"] and not plan_item.get("is_calculated"):
                errors.append(_error("source_field", f"{source_sheet}.{source_field}", "source field not found"))
                items.append(_empty_item(index, plan_item, "source_field_missing"))
                continue

            if write_type == "single_value":
                raw_value = _extract_single_value(sheet_data, source_field)
                cleaned = clean_value(raw_value)
                if cleaned["is_empty"] and plan_item.get("required"):
                    errors.append(_error("empty_value", f"{source_sheet}.{source_field}", "required value is empty"))
                items.append(
                    {
                        "plan_index": index,
                        "write_type": write_type,
                        "source_sheet": source_sheet,
                        "source_field": source_field,
                        "target_template": plan_item.get("target_template"),
                        "output_name": plan_item.get("output_name"),
                        "target_sheet": plan_item.get("target_sheet"),
                        "target_cell": plan_item.get("target_cell"),
                        "overwrite_formula": bool(plan_item.get("overwrite_formula")),
                        "required": bool(plan_item.get("required")),
                        "raw_value": cleaned["original_value"],
                        "cleaned_value": cleaned["cleaned_value"],
                        "is_empty": cleaned["is_empty"],
                        "is_calculated": False,
                    }
                )
            elif write_type == "detail_column":
                rows = []
                if not plan_item.get("is_calculated"):
                    for row in sheet_data["rows"]:
                        cleaned = clean_value(row["values"].get(source_field))
                        rows.append(
                            {
                                "source_row": row["row_number"],
                                "raw_value": cleaned["original_value"],
                                "cleaned_value": cleaned["cleaned_value"],
                                "is_empty": cleaned["is_empty"],
                            }
                        )
                items.append(
                    {
                        "plan_index": index,
                        "write_type": write_type,
                        "source_sheet": source_sheet,
                        "source_field": source_field,
                        "target_template": plan_item.get("target_template"),
                        "output_name": plan_item.get("output_name"),
                        "target_sheet": plan_item.get("target_sheet"),
                        "target_cell": plan_item.get("target_cell"),
                        "target_range": plan_item.get("target_range"),
                        "overwrite_formula": bool(plan_item.get("overwrite_formula")),
                        "required": bool(plan_item.get("required", True)),
                        "rows": rows,
                        "is_calculated": bool(plan_item.get("is_calculated")),
                        "calculation_formula": plan_item.get("calculation_formula", ""),
                    }
                )
    finally:
        workbook.close()

    return {
        "source_file": str(main_excel),
        "items": items,
        "sheet_rows": {
            sheet_name: {
                "header_row": data["header_row"],
                "fields": data["fields"],
                "rows": [
                    {
                        "row_number": row["row_number"],
                        "values": row["values"],
                        "cleaned_values": {field: value["cleaned_value"] for field, value in clean_row(row["values"]).items()},
                    }
                    for row in data["rows"]
                ],
            }
            for sheet_name, data in sheet_cache.items()
        },
    }, errors


def _get_sheet_data(workbook: Any, sheet_name: str, cache: dict[str, dict[str, Any]], errors: list[dict[str, str]]) -> dict[str, Any] | None:
    if sheet_name in cache:
        return cache[sheet_name]
    if sheet_name not in workbook.sheetnames:
        errors.append(_error("source_sheet", sheet_name, "source sheet not found"))
        return None

    worksheet = workbook[sheet_name]
    header_row, fields = detect_header_row(worksheet)
    if header_row is None:
        errors.append(_error("header", sheet_name, "header row not detected"))
        return None

    field_to_column = {field: index + 1 for index, field in enumerate(fields)}
    rows: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, (worksheet.max_row or 0) + 1):
        values = {
            field: worksheet.cell(row=row_index, column=column_index).value
            for field, column_index in field_to_column.items()
        }
        if all(value is None for value in values.values()):
            continue
        rows.append({"row_number": row_index, "values": values})

    cache[sheet_name] = {
        "header_row": header_row,
        "fields": fields,
        "rows": rows,
    }
    return cache[sheet_name]


def _extract_single_value(sheet_data: dict[str, Any], source_field: str) -> Any:
    for row in sheet_data["rows"]:
        value = row["values"].get(source_field)
        if value is not None:
            return value
    return None


def _empty_item(index: int, plan_item: dict[str, Any], status: str) -> dict[str, Any]:
    return {
        "plan_index": index,
        "write_type": plan_item.get("write_type"),
        "source_sheet": plan_item.get("source_sheet"),
        "source_field": plan_item.get("source_field"),
        "target_template": plan_item.get("target_template"),
        "output_name": plan_item.get("output_name"),
        "target_sheet": plan_item.get("target_sheet"),
        "target_cell": plan_item.get("target_cell"),
        "overwrite_formula": bool(plan_item.get("overwrite_formula")),
        "required": bool(plan_item.get("required")),
        "status": status,
    }


def _error(category: str, location: str, message: str) -> dict[str, str]:
    return {"category": category, "location": location, "message": message}
