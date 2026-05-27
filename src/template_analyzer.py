"""Analyze template workbook structure without modifying the workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def analyze_template(path: str | Path) -> dict[str, Any]:
    template_path = Path(path)
    workbook = load_workbook(template_path, read_only=False, data_only=False, keep_vba=template_path.suffix.lower() == ".xlsm")
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            sheets.append(
                {
                    "sheet_name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "merged_cells": [str(range_ref) for range_ref in worksheet.merged_cells.ranges],
                    "formula_cells": _formula_cells(worksheet),
                    "non_empty_cells": _non_empty_cells(worksheet),
                    "possible_detail_areas": _possible_detail_areas(worksheet),
                }
            )
        return {
            "file_name": template_path.name,
            "file_path": str(template_path),
            "sheet_names": workbook.sheetnames,
            "sheets": sheets,
        }
    finally:
        workbook.close()


def analyze_templates(paths: list[str | Path]) -> dict[str, Any]:
    templates = [analyze_template(path) for path in paths]
    return {"templates": templates}


def _formula_cells(worksheet: Any) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cells.append({"cell": cell.coordinate, "formula": cell.value})
    return cells


def _non_empty_cells(worksheet: Any) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append({"cell": cell.coordinate, "value_type": type(cell.value).__name__})
    return cells


def _possible_detail_areas(worksheet: Any) -> list[dict[str, Any]]:
    areas: list[dict[str, Any]] = []
    for row_index in range(1, (worksheet.max_row or 0) + 1):
        row_values = [worksheet.cell(row=row_index, column=col).value for col in range(1, (worksheet.max_column or 0) + 1)]
        non_empty_columns = [
            col_index
            for col_index, value in enumerate(row_values, start=1)
            if value is not None and str(value).strip()
        ]
        if len(non_empty_columns) < 2:
            continue

        next_row_index = row_index + 1
        if next_row_index > (worksheet.max_row or 0):
            continue
        next_has_space = any(
            worksheet.cell(row=next_row_index, column=col_index).value is None
            for col_index in non_empty_columns
        )
        if next_has_space:
            areas.append(
                {
                    "header_row": row_index,
                    "data_start_row": next_row_index,
                    "start_column": min(non_empty_columns),
                    "end_column": max(non_empty_columns),
                    "reason": "row has multiple labels and the next row has writable empty cells",
                }
            )
    return areas
