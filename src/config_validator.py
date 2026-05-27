"""Validate LLM-generated (or human-written) mapping config drafts.

Checks every field against actual workbook structures and reports issues
without modifying any Excel files.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_to_tuple

from src.rule_parser import MappingConfig, load_mapping_config


NEED_CONFIRM_MARKER = "NEED_CONFIRM"


def validate_draft_config(
    draft: dict[str, Any],
    main_structure: dict[str, Any] | None,
    template_structure: dict[str, Any] | None,
) -> tuple[list[dict[str, Any]], list[dict[str, str]], list[dict[str, Any]]]:
    """Validate a draft mapping config against workbook structures.

    Returns (passed, need_confirm, failed) lists.
    """
    passed: list[dict[str, Any]] = []
    need_confirm: list[dict[str, str]] = []
    failed: list[dict[str, Any]] = []

    _check_json_structure(draft, failed)
    _check_need_confirm_placeholders(draft, need_confirm)

    if not draft.get("templates"):
        failed.append(_issue("structure", "templates", "no templates defined"))
        return passed, need_confirm, failed

    main_sheets = _sheets_by_name(main_structure)
    templates = _templates_by_name(template_structure)

    _check_main_file(draft, main_structure, passed, failed)
    _check_source_sheets(draft, main_sheets, passed, need_confirm, failed)
    _check_template_names(draft, templates, passed, need_confirm, failed)
    _check_target_sheets(draft, templates, passed, need_confirm, failed)
    _check_mappings(draft, main_sheets, templates, passed, need_confirm, failed)
    _check_detail_areas(draft, main_sheets, templates, passed, need_confirm, failed)
    _check_calculations(draft, main_sheets, passed, need_confirm, failed)

    return passed, need_confirm, failed


def has_unresolved_confirmations(draft: dict[str, Any]) -> bool:
    """Check recursively whether any value starts with NEED_CONFIRM."""
    return _scan_need_confirm(draft)


def is_ready_for_fill(
    draft: dict[str, Any],
    main_structure: dict[str, Any] | None,
    template_structure: dict[str, Any] | None,
) -> tuple[bool, str]:
    """Return (ready, reason)."""
    _, need_confirm, failed = validate_draft_config(
        draft, main_structure, template_structure
    )
    if failed:
        return False, f"{len(failed)} validation failure(s)"
    if has_unresolved_confirmations(draft):
        return False, "unresolved NEED_CONFIRM fields remain"
    if need_confirm:
        return False, f"{len(need_confirm)} field(s) require confirmation"
    return True, "ready"


def write_validation_report(
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
    path: str | Path,
) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    for sheet_name, rows, headers in [
        ("passed", passed, ["check", "location", "message"]),
        ("need_confirm", need_confirm, ["field", "location", "message"]),
        ("failed", failed, ["check", "location", "message"]),
    ]:
        if rows:
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(headers)
            for row in rows:
                sheet.append([_str(row.get(h, "")) for h in headers])
    # Remove default sheet if unused
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]
    if workbook.sheetnames:
        workbook.save(output_path)


# ---- internal check functions ----


def _check_json_structure(draft: dict[str, Any], failed: list[dict[str, Any]]) -> None:
    try:
        MappingConfig.model_validate(draft)
    except Exception as exc:
        failed.append(_issue("json_schema", "root", f"config fails schema validation: {exc}"))


def _check_need_confirm_placeholders(
    draft: dict[str, Any], need_confirm: list[dict[str, str]]
) -> None:
    for path, value in _walk_need_confirm(draft, ""):
        need_confirm.append(
            {"field": path, "location": path, "message": f"unresolved placeholder: {value}"}
        )


def _check_main_file(
    draft: dict[str, Any],
    main_structure: dict[str, Any] | None,
    passed: list[dict[str, Any]],
    failed: list[dict[str, Any]],
) -> None:
    config_main = draft.get("main_file", "")
    if config_main.startswith(NEED_CONFIRM_MARKER):
        return
    if main_structure is None:
        failed.append(_issue("main_file", config_main, "main file structure not available for cross-check"))
        return
    actual_name = main_structure.get("file_name", "")
    if config_main != actual_name:
        failed.append(
            _issue("main_file", config_main, f"configured main_file does not match detected file '{actual_name}'")
        )
    else:
        passed.append(_issue("main_file", config_main, "main_file matches detected file"))


def _check_source_sheets(
    draft: dict[str, Any],
    main_sheets: dict[str, dict[str, Any]],
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
) -> None:
    for template_cfg in draft.get("templates", []):
        sheet = template_cfg.get("source_sheet", "")
        if str(sheet).startswith(NEED_CONFIRM_MARKER):
            need_confirm.append(
                {"field": "source_sheet", "location": str(sheet), "message": "source_sheet is unresolved"}
            )
            continue
        if sheet in main_sheets:
            passed.append(_issue("source_sheet", sheet, "source_sheet exists in main Excel"))
        else:
            failed.append(
                _issue(
                    "source_sheet",
                    sheet,
                    f"source_sheet not found in main Excel; available: {list(main_sheets.keys())}",
                )
            )

        detail = template_cfg.get("detail_area")
        if isinstance(detail, dict):
            detail_sheet = detail.get("source_sheet", "")
            if str(detail_sheet).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append(
                    {"field": "detail_area.source_sheet", "location": str(detail_sheet), "message": "detail_area source_sheet is unresolved"}
                )
            elif detail_sheet and detail_sheet not in main_sheets:
                failed.append(
                    _issue("detail_area.source_sheet", detail_sheet, "detail_area source_sheet not found in main Excel")
                )


def _check_template_names(
    draft: dict[str, Any],
    templates: dict[str, dict[str, Any]],
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
) -> None:
    available = list(templates.keys())
    for template_cfg in draft.get("templates", []):
        name = template_cfg.get("template_name", "")
        if str(name).startswith(NEED_CONFIRM_MARKER):
            need_confirm.append(
                {"field": "template_name", "location": str(name), "message": "template_name is unresolved"}
            )
            continue
        if name in templates:
            passed.append(_issue("template_name", name, "template exists"))
        else:
            failed.append(
                _issue("template_name", name, f"template not found; available: {available}")
            )


def _check_target_sheets(
    draft: dict[str, Any],
    templates: dict[str, dict[str, Any]],
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
) -> None:
    for template_cfg in draft.get("templates", []):
        tpl_name = template_cfg.get("template_name", "")
        tpl_struct = templates.get(tpl_name, {})
        tpl_sheets = _sheets_by_name(tpl_struct)

        target = template_cfg.get("target_sheet", "")
        if str(target).startswith(NEED_CONFIRM_MARKER):
            need_confirm.append(
                {"field": "target_sheet", "location": f"{tpl_name}:{target}", "message": "target_sheet is unresolved"}
            )
            continue
        if tpl_struct:
            if target in tpl_sheets:
                passed.append(_issue("target_sheet", f"{tpl_name}:{target}", "target_sheet exists in template"))
            else:
                failed.append(
                    _issue(
                        "target_sheet",
                        f"{tpl_name}:{target}",
                        f"target_sheet not found in template; available: {list(tpl_sheets.keys())}",
                    )
                )

        detail = template_cfg.get("detail_area")
        if isinstance(detail, dict):
            detail_target = detail.get("target_sheet", "")
            if str(detail_target).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append(
                    {"field": "detail_area.target_sheet", "location": str(detail_target), "message": "detail_area target_sheet is unresolved"}
                )
            elif tpl_struct and detail_target and detail_target not in tpl_sheets:
                failed.append(
                    _issue(
                        "detail_area.target_sheet",
                        f"{tpl_name}:{detail_target}",
                        "detail_area target_sheet not found in template",
                    )
                )


def _check_mappings(
    draft: dict[str, Any],
    main_sheets: dict[str, dict[str, Any]],
    templates: dict[str, dict[str, Any]],
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
) -> None:
    for template_cfg in draft.get("templates", []):
        tpl_name = template_cfg.get("template_name", "")
        tpl_struct = templates.get(tpl_name, {})
        tpl_sheets = _sheets_by_name(tpl_struct)
        target_sheet_name = template_cfg.get("target_sheet", "")
        target_sheet = tpl_sheets.get(target_sheet_name, {})
        source_sheet_name = template_cfg.get("source_sheet", "")
        source_sheet = main_sheets.get(source_sheet_name, {})
        source_fields = set(source_sheet.get("fields", []))
        formula_cell_refs = {
            item["cell"] for item in (target_sheet.get("formula_cells") or [])
            if isinstance(item, dict) and "cell" in item
        }

        for mapping in template_cfg.get("mappings", []):
            loc = f"{tpl_name}:{target_sheet_name}!{mapping.get('target_cell', '')}"

            source_field = mapping.get("source_field", "")
            if str(source_field).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append({"field": "source_field", "location": loc, "message": f"source_field is unresolved: {source_field}"})
            elif source_field and source_sheet and source_field not in source_fields:
                failed.append(_issue("source_field", loc, f"source_field '{source_field}' not found; available: {sorted(source_fields)}"))

            target_cell = mapping.get("target_cell", "")
            if str(target_cell).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append({"field": "target_cell", "location": loc, "message": f"target_cell is unresolved: {target_cell}"})
            elif target_cell:
                if not _is_valid_cell(target_cell):
                    failed.append(_issue("target_cell", loc, f"'{target_cell}' is not a valid cell reference"))
                elif target_sheet:
                    if target_cell in formula_cell_refs and not mapping.get("overwrite_formula"):
                        need_confirm.append(
                            {
                                "field": "overwrite_formula",
                                "location": loc,
                                "message": f"target cell contains a formula; set overwrite_formula=true to overwrite",
                            }
                        )
                    passed.append(_issue("mapping", loc, "single-value mapping validated"))


def _check_detail_areas(
    draft: dict[str, Any],
    main_sheets: dict[str, dict[str, Any]],
    templates: dict[str, dict[str, Any]],
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
) -> None:
    for template_cfg in draft.get("templates", []):
        detail = template_cfg.get("detail_area")
        if not isinstance(detail, dict):
            continue

        tpl_name = template_cfg.get("template_name", "")
        tpl_struct = templates.get(tpl_name, {})
        detail_sheet_name = detail.get("target_sheet", template_cfg.get("target_sheet", ""))
        detail_sheet = _sheets_by_name(tpl_struct).get(detail_sheet_name, {})
        max_row = detail_sheet.get("max_row", 0)

        start_row = detail.get("target_start_row")
        end_row = detail.get("target_end_row")

        if isinstance(start_row, str) and start_row.startswith(NEED_CONFIRM_MARKER):
            need_confirm.append({"field": "detail_area.target_start_row", "location": str(start_row), "message": "start row is unresolved"})
        if isinstance(end_row, str) and end_row.startswith(NEED_CONFIRM_MARKER):
            need_confirm.append({"field": "detail_area.target_end_row", "location": str(end_row), "message": "end row is unresolved"})

        if isinstance(start_row, int) and isinstance(end_row, int):
            if start_row > end_row:
                failed.append(
                    _issue("detail_area.range", f"{tpl_name}:{detail_sheet_name}", f"start_row ({start_row}) > end_row ({end_row})")
                )
            elif detail_sheet and end_row > max_row:
                failed.append(
                    _issue(
                        "detail_area.range",
                        f"{tpl_name}:{detail_sheet_name}",
                        f"end_row ({end_row}) exceeds template max_row ({max_row})",
                    )
                )
            else:
                passed.append(
                    _issue("detail_area.range", f"{tpl_name}:{detail_sheet_name}", f"rows {start_row}-{end_row} within bounds")
                )

        source_sheet_name = detail.get("source_sheet", template_cfg.get("source_sheet", ""))
        source_sheet = main_sheets.get(source_sheet_name, {})
        source_fields = set(source_sheet.get("fields", []))
        calculations = {calc.get("target_field", ""): calc.get("formula", "") for calc in detail.get("calculations", [])}

        for field, column in detail.get("columns", {}).items():
            loc = f"{tpl_name}:{detail_sheet_name}:{column}"
            if str(field).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append({"field": "detail_area.columns.field", "location": loc, "message": f"column field is unresolved: {field}"})
            elif field and field not in source_fields and field not in calculations:
                failed.append(_issue("detail_area.columns", loc, f"field '{field}' not found in source sheet"))
            if str(column).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append({"field": "detail_area.columns.column", "location": loc, "message": f"column letter is unresolved: {column}"})
            elif column and not _is_valid_detail_column(column):
                failed.append(_issue("detail_area.columns", loc, f"'{column}' is not a valid column letter"))


def _check_calculations(
    draft: dict[str, Any],
    main_sheets: dict[str, dict[str, Any]],
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
) -> None:
    for template_cfg in draft.get("templates", []):
        detail = template_cfg.get("detail_area")
        if not isinstance(detail, dict):
            continue
        source_sheet_name = detail.get("source_sheet", template_cfg.get("source_sheet", ""))
        source_sheet = main_sheets.get(source_sheet_name, {})
        source_fields = set(source_sheet.get("fields", []))
        detail_columns = set(detail.get("columns", {}).keys())

        for calc in detail.get("calculations", []):
            target_field = calc.get("target_field", "")
            formula = calc.get("formula", "")
            loc = f"calculation.{target_field}"

            if str(target_field).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append({"field": "calculation.target_field", "location": loc, "message": "target_field is unresolved"})
            elif target_field not in detail_columns:
                need_confirm.append(
                    {"field": "calculation.target_field", "location": loc, "message": f"'{target_field}' missing from detail_area.columns; add it"}
                )

            if str(formula).startswith(NEED_CONFIRM_MARKER):
                need_confirm.append({"field": "calculation.formula", "location": loc, "message": "formula is unresolved"})
            elif formula:
                referenced = _extract_formula_fields(formula)
                for ref in referenced:
                    if ref not in source_fields and ref not in detail_columns:
                        failed.append(
                            _issue("calculation.formula", loc, f"formula references unknown field '{ref}'; available: {sorted(source_fields | detail_columns)}")
                        )
                passed.append(_issue("calculation", loc, f"'{target_field} = {formula}' validated"))


# ---- helpers ----


def _scan_need_confirm(value: Any) -> bool:
    if isinstance(value, str):
        return value.startswith(NEED_CONFIRM_MARKER)
    if isinstance(value, dict):
        return any(_scan_need_confirm(v) for v in value.values())
    if isinstance(value, list):
        return any(_scan_need_confirm(v) for v in value)
    return False


def _walk_need_confirm(value: Any, path: str) -> list[tuple[str, str]]:
    results: list[tuple[str, str]] = []
    if isinstance(value, str) and value.startswith(NEED_CONFIRM_MARKER):
        results.append((path, value))
    elif isinstance(value, dict):
        for key, child in value.items():
            results.extend(_walk_need_confirm(child, f"{path}.{key}" if path else key))
    elif isinstance(value, list):
        for idx, child in enumerate(value):
            results.extend(_walk_need_confirm(child, f"{path}[{idx}]"))
    return results


def _extract_formula_fields(formula: str) -> set[str]:
    """Extract field/variable names from a calculation formula string."""
    import re
    tokens = re.findall(r'[A-Za-z_一-鿿][A-Za-z0-9_一-鿿]*', formula)
    keywords = {"sum", "SUM", "if", "IF", "and", "AND", "or", "OR", "not", "NOT", "true", "false", "True", "False"}
    return {t for t in tokens if t not in keywords}


def _sheets_by_name(structure: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not structure:
        return {}
    return {sheet["sheet_name"]: sheet for sheet in structure.get("sheets", [])}


def _templates_by_name(structure: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not structure:
        return {}
    return {tpl["file_name"]: tpl for tpl in structure.get("templates", [])}


def _is_valid_cell(value: str) -> bool:
    try:
        row, column = coordinate_to_tuple(value)
        return row >= 1 and column >= 1
    except (ValueError, AttributeError):
        return False


def _is_valid_detail_column(value: str) -> bool:
    return bool(value) and value.isascii() and value.isalpha() and value.upper() == value


def _issue(check: str, location: str, message: str) -> dict[str, Any]:
    return {"check": check, "location": location, "message": message}


def _str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)
