from __future__ import annotations

import json
import zipfile
from html import escape
from pathlib import Path
from typing import Any, Callable

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response

from main import run_analyze, run_all, run_plan
from src.nl_rule_parser import apply_upload_context, parse_rules_text, write_rules_outputs

_HTTPX_AVAILABLE = False
try:
    import httpx  # noqa: F401

    _HTTPX_AVAILABLE = True
except ImportError:
    pass

ROOT = Path(__file__).resolve().parent
MAIN_DIR = ROOT / "input" / "main_excel"
TEMPLATE_DIR = ROOT / "input" / "templates"
CONFIG_DIR = ROOT / "config"
OUTPUT_DIR = ROOT / "output"
WEB_DIR = OUTPUT_DIR / "web"

app = FastAPI(title="Excel Template Filler")


# ===================================================================
# pages
# ===================================================================


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    status = request.query_params.get("status", "")
    status_type = request.query_params.get("type", "info")
    edit = request.query_params.get("edit", "")
    llm_config = _read_llm_config()
    llm_results = _read_llm_results()
    validation = _read_validation()
    draft_json = _read_draft_raw()  # show editor whenever draft exists
    return HTMLResponse(
        _render_page(
            status=status,
            status_type=status_type,
            summary=_read_summary(),
            errors=_read_error_summary(),
            llm_config=llm_config,
            llm_results=llm_results,
            llm_available=_HTTPX_AVAILABLE,
            validation=validation,
            draft_json=draft_json,
        )
    )


# ===================================================================
# LLM config
# ===================================================================


@app.post("/save-llm-config")
async def save_llm_config(
    model: str = Form("gpt-4o"),
    api_key: str = Form(""),
    base_url: str = Form("https://api.openai.com/v1"),
    temperature: str = Form("0.1"),
    max_tokens: str = Form("8192"),
) -> RedirectResponse:
    try:
        config = {
            "provider": "openai",
            "model": model.strip(),
            "api_key": api_key.strip(),
            "base_url": base_url.strip().rstrip("/"),
            "temperature": float(temperature),
            "max_tokens": int(max_tokens),
            "request_timeout": 120,
        }
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        (CONFIG_DIR / "llm_config.json").write_text(
            json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return _status_redirect("LLM 配置已保存。", "success")
    except Exception as exc:
        return _status_redirect(f"保存 LLM 配置失败：{exc}", "error")


# ===================================================================
# upload & parse
# ===================================================================


@app.post("/upload")
async def upload_files(
    main_excel: UploadFile = File(...),
    templates: list[UploadFile] = File(default=[]),
    rules_text: str = Form(...),
    use_llm: str = Form("false"),
) -> RedirectResponse:
    try:
        MAIN_DIR.mkdir(parents=True, exist_ok=True)
        TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        _clear_excel_inputs(MAIN_DIR)
        _clear_excel_inputs(TEMPLATE_DIR)

        main_name = _safe_name(main_excel.filename)
        main_path = MAIN_DIR / main_name
        await _save_upload(main_excel, main_path)

        template_names: list[str] = []
        for template in templates:
            if template.filename:
                template_name = _safe_name(template.filename)
                await _save_upload(template, TEMPLATE_DIR / template_name)
                template_names.append(template_name)

        (CONFIG_DIR / "rules_prompt.txt").write_text(rules_text, encoding="utf-8")

        if use_llm == "true":
            return await _handle_llm_upload(main_name, main_path, template_names, rules_text)

        # deterministic parser
        draft, confirmations = parse_rules_text(rules_text)
        template_sheets_map = {
            name: _read_workbook_sheet_names(TEMPLATE_DIR / name) for name in template_names
        }
        apply_upload_context(
            draft, confirmations,
            main_file=main_name, template_names=template_names,
            main_sheet_names=_read_workbook_sheet_names(main_path),
            template_sheets_map=template_sheets_map, rules_text=rules_text,
        )
        write_rules_outputs(draft, confirmations,
                            CONFIG_DIR / "mapping_config.json",
                            CONFIG_DIR / "mapping_config.need_confirm.json")

        if draft.get("need_confirm"):
            unresolved = _unresolved_fields(draft)
            return _status_redirect(
                "警告：文件已上传，但规则仍需确认：" + "、".join(unresolved) + "。请补充规则后重新上传。",
                "warning")
        return _status_redirect(
            f"成功：已上传 1 个主表和 {len(template_names)} 个模板，并已生成可执行配置。",
            "success")
    except Exception as exc:
        return _status_redirect(f"失败：上传或规则解析失败，原因：{exc}", "error")


async def _handle_llm_upload(
    main_name: str, main_path: Path, template_names: list[str], rules_text: str,
) -> RedirectResponse:
    if not _HTTPX_AVAILABLE:
        return _status_redirect("失败：LLM 模式需要 httpx 库，请执行: pip install httpx", "error")

    llm_config_path = CONFIG_DIR / "llm_config.json"
    if not llm_config_path.exists():
        return _status_redirect("失败：未找到 LLM 配置，请先在页面顶部配置 LLM 连接信息。", "error")

    run_analyze()

    from src.llm_rule_parser import load_llm_config, parse_rules_with_llm, write_llm_outputs

    main_structure = json.loads(
        (OUTPUT_DIR / "logs" / "main_excel_structure.json").read_text(encoding="utf-8"))
    template_structure = json.loads(
        (OUTPUT_DIR / "logs" / "template_structure.json").read_text(encoding="utf-8"))

    llm_config = load_llm_config(llm_config_path)
    draft, need_confirm, log_entries = parse_rules_with_llm(
        rules_text=rules_text, main_structure=main_structure,
        template_structure=template_structure, llm_config=llm_config)

    draft_path = CONFIG_DIR / "mapping_config.draft.json"
    confirm_path = CONFIG_DIR / "mapping_config.need_confirm_questions.json"
    log_path = OUTPUT_DIR / "logs" / "rule_parse_log.xlsx"
    write_llm_outputs(draft, need_confirm, log_entries, draft_path, confirm_path, log_path)

    apply_upload_context(
        draft, need_confirm,
        main_file=main_name, template_names=template_names,
        main_sheet_names=_read_workbook_sheet_names(main_path),
        template_sheets_map={
            name: _read_workbook_sheet_names(TEMPLATE_DIR / name) for name in template_names
        },
        rules_text=rules_text,
    )
    write_llm_outputs(draft, need_confirm, log_entries, draft_path, confirm_path, log_path)

    from src.config_validator import validate_draft_config, is_ready_for_fill
    passed, need_confirm_issues, failed = validate_draft_config(draft, main_structure, template_structure)
    ready, reason = is_ready_for_fill(draft, main_structure, template_structure)

    # Save validation report
    _save_validation(passed, need_confirm_issues, failed, ready, reason)

    if ready:
        (CONFIG_DIR / "mapping_config.json").write_text(
            json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
        return _status_redirect(
            f"成功：LLM 已生成配置（{len(draft.get('templates', []))} 个模板），校验通过，可执行填充。",
            "success")
    else:
        return _status_redirect(
            f"警告：LLM 配置草稿已生成，{len(need_confirm)} 个待确认项。请在下方编辑器中修改草稿后重新校验。",
            "warning")


# ===================================================================
# draft editing & validation
# ===================================================================


@app.post("/save-draft")
async def save_draft(draft_text: str = Form(...)) -> RedirectResponse:
    """Save the user-edited draft JSON, then re-validate."""
    try:
        draft = json.loads(draft_text)
    except json.JSONDecodeError as exc:
        return _status_redirect(f"失败：JSON 格式错误 — {exc}", "error")

    try:
        draft_path = CONFIG_DIR / "mapping_config.draft.json"
        draft_path.write_text(json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")

        # Re-validate
        main_structure = _load_json_if_exists(OUTPUT_DIR / "logs" / "main_excel_structure.json")
        template_structure = _load_json_if_exists(OUTPUT_DIR / "logs" / "template_structure.json")

        from src.config_validator import validate_draft_config, is_ready_for_fill
        passed, need_confirm_issues, failed = validate_draft_config(
            draft, main_structure, template_structure)
        ready, reason = is_ready_for_fill(draft, main_structure, template_structure)
        _save_validation(passed, need_confirm_issues, failed, ready, reason)

        if ready:
            (CONFIG_DIR / "mapping_config.json").write_text(
                json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
            return _status_redirect(
                f"成功：草稿已保存，校验通过（{len(passed)} 项），已同步到正式配置，可执行填充。",
                "success")
        else:
            return _status_redirect(
                f"草稿已保存，但校验未通过：{reason}。{len(failed)} 项失败，{len(need_confirm_issues)} 项待确认。请继续修改。",
                "warning")
    except Exception as exc:
        return _status_redirect(f"失败：保存草稿时出错 — {exc}", "error")


@app.post("/promote-draft")
def promote_draft() -> RedirectResponse:
    """Copy the draft to mapping_config.json after final validation."""
    draft_path = CONFIG_DIR / "mapping_config.draft.json"
    if not draft_path.exists():
        return _status_redirect("失败：没有草稿可发布，请先运行 LLM 解析。", "error")

    try:
        draft = json.loads(draft_path.read_text(encoding="utf-8"))
        main_structure = _load_json_if_exists(OUTPUT_DIR / "logs" / "main_excel_structure.json")
        template_structure = _load_json_if_exists(OUTPUT_DIR / "logs" / "template_structure.json")

        from src.config_validator import is_ready_for_fill
        ready, reason = is_ready_for_fill(draft, main_structure, template_structure)
        if not ready:
            return _status_redirect(f"失败：草稿未通过校验 ({reason})，请先修改。", "error")

        (CONFIG_DIR / "mapping_config.json").write_text(
            json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
        return _status_redirect("成功：草稿已发布为正式配置，可以执行填充。", "success")
    except Exception as exc:
        return _status_redirect(f"失败：发布草稿时出错 — {exc}", "error")


@app.get("/view-draft-json", response_class=HTMLResponse)
def view_draft_json() -> HTMLResponse:
    """Full-page draft editor."""
    draft_json = _read_draft_raw()
    validation = _read_validation()
    return HTMLResponse(_render_draft_editor(draft_json, validation))


# ===================================================================
# actions
# ===================================================================


@app.post("/analyze")
def analyze() -> RedirectResponse:
    message, status_type = _run_action(
        run_analyze, "成功：文件结构分析完成。", "警告：分析完成但存在异常，请查看报告。")
    return _status_redirect(message, status_type)


@app.post("/plan")
def plan() -> RedirectResponse:
    config_path = CONFIG_DIR / "mapping_config.json"
    if not config_path.exists():
        draft_path = CONFIG_DIR / "mapping_config.draft.json"
        if draft_path.exists():
            return _status_redirect(
                "失败：尚未生成正式配置。请先在 LLM 解析结果中编辑草稿并通过校验，或点击「发布为正式配置」。",
                "error")
        return _status_redirect("失败：没有配置文件，请先上传文件并生成配置。", "error")

    message, status_type = _run_action(
        lambda: run_plan(config_path),
        "成功：填充计划已生成。",
        "警告：填充计划已生成但存在配置异常，请查看报告。")
    return _status_redirect(message, status_type)


@app.post("/fill")
async def fill(request: Request) -> RedirectResponse:
    form = await request.form()
    shutdown_after = form.get("shutdown", "false") == "true"

    config_path = CONFIG_DIR / "mapping_config.json"
    if not config_path.exists():
        draft_path = CONFIG_DIR / "mapping_config.draft.json"
        if draft_path.exists():
            return _status_redirect(
                "失败：尚未生成正式配置。请在 LLM 解析结果中编辑草稿并通过校验后，点击「发布为正式配置」。",
                "error")
        return _status_redirect("失败：没有配置文件，请先上传文件并生成配置。", "error")

    message, status_type = _run_action(
        lambda: run_all(config_path),
        "成功：Excel 已填充完成，可以下载修改后的文件。",
        "警告：流程已结束但存在异常，请先查看报告摘要。")

    if shutdown_after and status_type == "success":
        _schedule_shutdown()

    return _status_redirect(message, status_type)


# ===================================================================
# download
# ===================================================================


@app.get("/download")
def download_output() -> Response:
    WEB_DIR.mkdir(parents=True, exist_ok=True)
    zip_path = WEB_DIR / "filled_excel_files.zip"
    if zip_path.exists():
        zip_path.unlink()
    filled_dir = OUTPUT_DIR / "filled_files"
    excel_files = [
        path for path in sorted(filled_dir.glob("*"))
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
    ]
    if not excel_files:
        return _status_redirect("失败：没有找到已填充的 Excel 文件，请先执行填充。", "error")
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in excel_files:
            archive.write(path, path.name)
    return FileResponse(zip_path, filename="filled-excel-files.zip", media_type="application/zip")


@app.get("/download-llm-draft")
def download_llm_draft() -> Response:
    draft_path = CONFIG_DIR / "mapping_config.draft.json"
    if not draft_path.exists():
        return _status_redirect("失败：未找到 LLM 生成的草稿文件。", "error")
    return FileResponse(draft_path, filename="mapping_config.draft.json", media_type="application/json")


@app.get("/download-llm-confirm")
def download_llm_confirm() -> Response:
    confirm_path = CONFIG_DIR / "mapping_config.need_confirm_questions.json"
    if not confirm_path.exists():
        return _status_redirect("失败：未找到待确认问题文件。", "error")
    return FileResponse(confirm_path, filename="need_confirm_questions.json", media_type="application/json")


# ===================================================================
# helpers
# ===================================================================


async def _save_upload(upload: UploadFile, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("wb") as file:
        while chunk := await upload.read(1024 * 1024):
            file.write(chunk)


def _clear_excel_inputs(directory: Path) -> None:
    for path in directory.iterdir() if directory.exists() else []:
        if path.name == ".gitkeep":
            continue
        if path.is_file():
            path.unlink()


def _safe_name(filename: str | None) -> str:
    name = Path(filename or "upload.xlsx").name
    return name.replace("/", "_").replace("\\", "_")


def _read_workbook_sheet_names(path: Path) -> list[str]:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return []
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=False,
                                 keep_vba=path.suffix.lower() == ".xlsm")
        names = list(workbook.sheetnames)
        workbook.close()
        return names
    except Exception:
        return []


def _status_redirect(message: str, status_type: str) -> RedirectResponse:
    from urllib.parse import urlencode
    query = urlencode({"status": message, "type": status_type})
    return RedirectResponse(f"/?{query}", status_code=303)


def _run_action(action: Callable[[], int], success_message: str, warning_message: str) -> tuple[str, str]:
    try:
        code = action()
        if code == 0:
            return success_message, "success"
        return warning_message, "warning"
    except Exception as exc:
        return f"失败：执行失败，原因：{exc}", "error"


def _read_summary() -> str:
    summary = OUTPUT_DIR / "summary.txt"
    if summary.exists():
        return summary.read_text(encoding="utf-8", errors="replace")
    return "暂无处理摘要。"


def _read_error_summary() -> str:
    report = OUTPUT_DIR / "reports" / "error_report.xlsx"
    if not report.exists():
        return "暂无异常报告。"
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(report, read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 8), values_only=True):
            if any(value is not None for value in row):
                rows.append(" | ".join("" if value is None else str(value) for value in row))
        workbook.close()
        return "\n".join(rows) if rows else "未记录异常。"
    except Exception as exc:
        return f"读取异常报告失败：{exc}"


def _read_llm_config() -> dict[str, Any]:
    llm_config_path = CONFIG_DIR / "llm_config.json"
    if llm_config_path.exists():
        try:
            return json.loads(llm_config_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _read_llm_results() -> dict[str, Any]:
    results: dict[str, Any] = {"has_draft": False, "has_confirm": False}
    draft_path = CONFIG_DIR / "mapping_config.draft.json"
    confirm_path = CONFIG_DIR / "mapping_config.need_confirm_questions.json"

    if draft_path.exists():
        try:
            draft = json.loads(draft_path.read_text(encoding="utf-8"))
            results["has_draft"] = True
            results["project_name"] = draft.get("project_name", "")
            results["main_file"] = draft.get("main_file", "")
            results["template_count"] = len(draft.get("templates", []))
            templates_info = []
            for tpl in draft.get("templates", []):
                templates_info.append({
                    "name": tpl.get("template_name", ""),
                    "source_sheet": tpl.get("source_sheet", ""),
                    "target_sheet": tpl.get("target_sheet", ""),
                    "mapping_count": len(tpl.get("mappings", [])),
                    "has_detail": "detail_area" in tpl,
                })
            results["templates"] = templates_info
            results["need_confirm_flag"] = draft.get("need_confirm", False)
            results["unresolved_count"] = len(_unresolved_fields(draft))
        except Exception:
            pass

    if confirm_path.exists():
        try:
            confirm_data = json.loads(confirm_path.read_text(encoding="utf-8"))
            items = confirm_data.get("need_confirm", [])
            results["has_confirm"] = True
            results["confirm_items"] = items[:30]
            results["confirm_total"] = len(items)
        except Exception:
            pass

    return results


def _read_validation() -> dict[str, Any]:
    vp = CONFIG_DIR / "validation_result.json"
    if vp.exists():
        try:
            return json.loads(vp.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_validation(
    passed: list[dict[str, Any]],
    need_confirm: list[dict[str, str]],
    failed: list[dict[str, Any]],
    ready: bool,
    reason: str,
) -> None:
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    (CONFIG_DIR / "validation_result.json").write_text(json.dumps({
        "passed": len(passed), "need_confirm": len(need_confirm),
        "failed": len(failed), "ready": ready, "reason": reason,
        "failed_items": [f"{f['check']}: {f['message']}" for f in failed[:20]],
        "need_confirm_items": [
            f"{n.get('field','')}: {n.get('message','')}" for n in need_confirm[:20]
        ],
    }, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_draft_raw() -> str:
    draft_path = CONFIG_DIR / "mapping_config.draft.json"
    if draft_path.exists():
        try:
            return draft_path.read_text(encoding="utf-8")
        except Exception:
            pass
    return ""


def _schedule_shutdown() -> None:
    """Schedule a system shutdown with 60-second delay (Windows only)."""
    import subprocess
    try:
        subprocess.Popen(["shutdown", "/s", "/t", "60", "/c", "Excel Template Filler 已完成，系统将在 60 秒后关机。运行 shutdown /a 取消。"])
    except Exception:
        pass  # best-effort


def _load_json_if_exists(path: Path) -> Any:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return None


def _unresolved_fields(value: object) -> list[str]:
    fields: list[str] = []

    def walk(item: object, path: str) -> None:
        if isinstance(item, str) and item.startswith("NEED_CONFIRM"):
            fields.append(path)
        elif isinstance(item, dict):
            for key, child in item.items():
                walk(child, key)
        elif isinstance(item, list):
            for child in item:
                walk(child, path)

    walk(value, "配置")
    return sorted(set(fields)) or ["配置"]


# ===================================================================
# HTML renderer
# ===================================================================


def _render_page(
    status: str,
    status_type: str,
    summary: str,
    errors: str,
    llm_config: dict[str, Any],
    llm_results: dict[str, Any],
    llm_available: bool,
    validation: dict[str, Any],
    draft_json: str,
) -> str:
    safe_status = escape(status)
    safe_status_type = escape(status_type if status_type in {"success", "warning", "error", "info"} else "info")
    safe_summary = escape(summary)
    safe_errors = escape(errors)

    cfg_model = escape(str(llm_config.get("model", "gpt-4o")))
    cfg_key = escape(str(llm_config.get("api_key", "")))
    cfg_url = escape(str(llm_config.get("base_url", "https://api.openai.com/v1")))
    cfg_temp = escape(str(llm_config.get("temperature", "0.1")))
    cfg_tokens = escape(str(llm_config.get("max_tokens", "8192")))
    cfg_configured = "已配置" if llm_config.get("api_key") else "未配置"
    cfg_status_class = "success" if llm_config.get("api_key") else "warning"

    llm_disabled = "" if llm_available else "disabled"
    llm_hint = "" if llm_available else "（需安装 httpx: pip install httpx）"

    # LLM results summary
    llm_section = _render_llm_section(llm_results)
    # Draft editor
    editor_section = _render_editor_section(draft_json, validation, llm_results)
    # Validation section
    validation_section = _render_validation_section(validation)

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Excel Template Filler</title>
  <style>
    :root {{
      --bg: #f6f7f9; --text: #18202a; --muted: #657080; --line: #d9dee7;
      --surface: #ffffff; --accent: #2166d1; --accent-dark: #174c9f;
      --ok: #0f7a46; --warn: #936100; --bad: #b42318; --info: #36516d;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0; font-family: Arial, "Microsoft YaHei", sans-serif;
      background: var(--bg); color: var(--text);
    }}
    header {{
      background: var(--surface); border-bottom: 1px solid var(--line);
      padding: 18px 28px; display: flex; align-items: center;
      justify-content: space-between; gap: 16px;
    }}
    h1 {{ margin: 0; font-size: 22px; }}
    main {{ max-width: 1120px; margin: 28px auto; padding: 0 20px 40px; }}
    section {{
      background: var(--surface); border: 1px solid var(--line);
      border-radius: 8px; padding: 20px; margin-bottom: 18px;
    }}
    h2 {{ margin: 0 0 16px; font-size: 17px; }}
    h3 {{ margin: 14px 0 8px; font-size: 14px; color: var(--muted); }}
    label {{ display: block; margin: 12px 0 6px; font-size: 13px; color: var(--muted); }}
    input[type="file"], textarea, input[type="text"], input[type="password"], input[type="number"] {{
      width: 100%; padding: 10px; border: 1px solid var(--line); border-radius: 6px;
      background: #fbfcfd; font-family: inherit; font-size: 14px; line-height: 1.5;
    }}
    textarea {{ min-height: 150px; resize: vertical; font-family: "Cascadia Code", "Consolas", monospace; font-size: 13px; }}
    .draft-editor {{ min-height: 350px; }}
    .actions {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }}
    button, .download {{
      border: 0; background: var(--accent); color: white; padding: 10px 14px;
      border-radius: 6px; font-size: 14px; line-height: 1; cursor: pointer;
      text-decoration: none; display: inline-flex; align-items: center; min-height: 38px;
    }}
    button:hover, .download:hover {{ background: var(--accent-dark); }}
    button:disabled {{ background: #a0acba; cursor: not-allowed; }}
    .secondary {{ background: #2f3a4a; }}
    .ok-btn {{ background: var(--ok); }}
    .ok-btn:hover {{ background: #0b5e35; }}
    .warn-btn {{ background: var(--warn); }}
    .warn-btn:hover {{ background: #7a5000; }}
    .status {{
      font-size: 14px; min-height: 20px; white-space: normal; overflow-wrap: anywhere;
      padding: 8px 10px; border-radius: 6px; border: 1px solid transparent;
      background: #eef3f8; color: var(--info);
    }}
    .status.success {{ color: var(--ok); background: #edf8f2; border-color: #bde5ce; }}
    .status.warning {{ color: var(--warn); background: #fff8e6; border-color: #f2d68a; }}
    .status.error {{ color: var(--bad); background: #fff1f0; border-color: #f3b7b2; }}
    .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }}
    .grid-3 {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
    pre {{
      margin: 0; padding: 14px; border-radius: 6px;
      border: 1px solid var(--line); background: #fbfcfd;
      white-space: pre-wrap; overflow-wrap: anywhere; min-height: 140px;
      font-size: 13px; line-height: 1.5; color: #263241;
    }}
    .badge {{
      display: inline-block; padding: 2px 8px; border-radius: 10px;
      font-size: 12px; font-weight: 600;
    }}
    .badge.ok {{ background: #edf8f2; color: var(--ok); }}
    .badge.warn {{ background: #fff8e6; color: var(--warn); }}
    .badge.err {{ background: #fff1f0; color: var(--bad); }}
    details {{ margin-top: 8px; }}
    summary {{ cursor: pointer; font-size: 14px; color: var(--accent); padding: 4px 0; }}
    summary:hover {{ color: var(--accent-dark); }}
    .llm-config-row {{ display: flex; gap: 10px; align-items: flex-end; flex-wrap: wrap; }}
    .llm-config-row .field {{ flex: 1; min-width: 140px; }}
    .result-table, .val-table {{
      width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px;
    }}
    .result-table th, .result-table td, .val-table th, .val-table td {{
      text-align: left; padding: 6px 8px; border-bottom: 1px solid var(--line);
    }}
    .result-table th, .val-table th {{ color: var(--muted); font-weight: 600; font-size: 12px; }}
    .config-guide {{
      font-size: 13px; line-height: 1.7; color: var(--muted);
    }}
    .config-guide code {{
      background: #eef3f8; padding: 1px 5px; border-radius: 3px; font-size: 12px;
    }}
    .config-guide table {{ width: 100%; border-collapse: collapse; margin: 8px 0; font-size: 12px; }}
    .config-guide th, .config-guide td {{
      text-align: left; padding: 6px 8px; border-bottom: 1px solid var(--line);
    }}
    .config-guide th {{ color: var(--text); font-weight: 600; }}
    .inline-form {{ display: inline; }}
    .warning-box {{
      background: #fff8e6; border: 1px solid #f2d68a; border-radius: 6px;
      padding: 12px 16px; margin: 12px 0; font-size: 13px; color: var(--warn);
    }}
    .val-summary {{
      display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 12px;
    }}
    .val-chip {{
      padding: 6px 14px; border-radius: 6px; font-size: 14px; font-weight: 600;
      text-align: center; min-width: 80px;
    }}
    .val-chip.pass {{ background: #edf8f2; color: var(--ok); }}
    .val-chip.confirm {{ background: #fff8e6; color: var(--warn); }}
    .val-chip.fail {{ background: #fff1f0; color: var(--bad); }}
    @media (max-width: 760px) {{
      header {{ flex-direction: column; align-items: flex-start; }}
      .grid, .grid-3, .grid-2 {{ grid-template-columns: 1fr; }}
      main {{ margin-top: 18px; padding-inline: 14px; }}
      .llm-config-row {{ flex-direction: column; }}
      .val-summary {{ flex-direction: column; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>Excel 多模板自动填充系统</h1>
    <div class="status {safe_status_type}">{safe_status or "等待操作：请先上传文件并输入自然语言规则。"}</div>
  </header>
  <main>
    <!-- LLM Configuration -->
    <section>
      <h2>LLM 配置 <span class="badge {cfg_status_class}">{cfg_configured}</span></h2>
      <form action="/save-llm-config" method="post">
        <div class="llm-config-row">
          <div class="field">
            <label>模型名称</label>
            <input type="text" name="model" value="{cfg_model}" placeholder="gpt-4o / deepseek-chat">
          </div>
          <div class="field">
            <label>API Key</label>
            <input type="password" name="api_key" value="{cfg_key}" placeholder="sk-...">
          </div>
          <div class="field">
            <label>API 地址</label>
            <input type="text" name="base_url" value="{cfg_url}" placeholder="https://api.openai.com/v1">
          </div>
        </div>
        <div class="grid-2" style="margin-top:10px">
          <div><label>Temperature (0-2)</label><input type="number" name="temperature" value="{cfg_temp}" min="0" max="2" step="0.1"></div>
          <div><label>Max Tokens</label><input type="number" name="max_tokens" value="{cfg_tokens}" min="256" max="131072" step="1"></div>
        </div>
        <div style="margin-top:14px"><button type="submit">保存 LLM 配置</button></div>
      </form>
      <details>
        <summary>配置说明与常用服务商</summary>
        <div class="config-guide">
          <table>
            <tr><th>服务商</th><th>base_url</th><th>常用模型</th></tr>
            <tr><td>OpenAI</td><td><code>https://api.openai.com/v1</code></td><td>gpt-4o, gpt-4o-mini</td></tr>
            <tr><td>DeepSeek</td><td><code>https://api.deepseek.com/v1</code></td><td>deepseek-chat, deepseek-reasoner</td></tr>
            <tr><td>通义千问</td><td><code>https://dashscope.aliyuncs.com/compatible-mode/v1</code></td><td>qwen-plus, qwen-max</td></tr>
            <tr><td>Ollama (本地)</td><td><code>http://localhost:11434/v1</code></td><td>llama3, qwen2.5 等</td></tr>
            <tr><td>智谱 GLM</td><td><code>https://open.bigmodel.cn/api/paas/v4</code></td><td>glm-4-plus, glm-4-flash</td></tr>
            <tr><td>Moonshot</td><td><code>https://api.moonshot.cn/v1</code></td><td>moonshot-v1-8k</td></tr>
          </table>
          <p><strong>使用流程：</strong> 配置 LLM → 上传文件并勾选「使用 LLM」→ 编辑草稿解决待确认项 → 校验通过后发布 → 执行填充</p>
          <p><strong>安全约束：</strong>LLM 只生成 JSON 配置文件，<strong>不直接修改任何 Excel 文件</strong>。不确定的信息标记为 <code>NEED_CONFIRM_xxx</code>。</p>
        </div>
      </details>
    </section>

    <!-- Upload -->
    <section>
      <h2>上传文件</h2>
      <form action="/upload" method="post" enctype="multipart/form-data">
        <label>主 Excel</label>
        <input type="file" name="main_excel" accept=".xlsx,.xlsm,.xls" required>
        <label>模板 Excel，可多选</label>
        <input type="file" name="templates" accept=".xlsx,.xlsm,.xls" multiple required>
        <label>自然语言规则（描述哪些字段填到模板的哪些位置）</label>
        <textarea name="rules_text" required placeholder="主表 Sheet「发票数据」填入 INVOICE.xlsx 模板。&#10;客户名称填入 B5。&#10;发票号填入 F3。&#10;品名、数量、单价、金额填入第 10 行到第 30 行。&#10;金额 = 数量 × 单价。"></textarea>
        <div style="margin-top:12px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">
          <label style="margin:0;display:flex;align-items:center;gap:6px;cursor:pointer;font-size:14px;color:var(--text)">
            <input type="checkbox" name="use_llm" value="true" {llm_disabled}>
            使用 LLM 解析规则 {llm_hint}
          </label>
          <button type="submit">上传文件</button>
        </div>
      </form>
    </section>

    <!-- LLM Results -->
    {llm_section}

    <!-- Validation -->
    {validation_section}

    <!-- Draft Editor -->
    {editor_section}

    <!-- Execute -->
    <section>
      <h2>执行流程</h2>
      <div class="actions">
        <form action="/analyze" method="post" class="inline-form"><button type="submit">分析文件</button></form>
        <form action="/plan" method="post" class="inline-form"><button type="submit">生成填充计划</button></form>
        <form action="/fill" method="post" class="inline-form" id="fill-form">
          <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
            <button type="submit">执行填充</button>
            <label style="margin:0;display:flex;align-items:center;gap:4px;cursor:pointer;font-size:13px;color:var(--text)">
              <input type="checkbox" name="shutdown" value="true">
              执行完成后关机
            </label>
          </div>
        </form>
        <a class="download secondary" href="/download">下载已填充 Excel</a>
      </div>
      <details style="margin-top:12px">
        <summary>重要提示</summary>
        <div class="warning-box">
          <strong>已知限制：</strong>openpyxl 库在保存 Excel 时不保留模板中的<strong>图片、形状、公章</strong>等对象。
          如果模板含有公章图片，填充后的文件会丢失这些图片。建议在填充完成后，手动将公章图片重新插入到输出文件中。
          如果需要自动保留图片，可以使用 xlwings（需安装 Excel）替代 openpyxl。
        </div>
      </details>
    </section>

    <!-- Summary -->
    <div class="grid">
      <section>
        <h2>异常报告摘要</h2>
        <pre>{safe_errors}</pre>
      </section>
      <section>
        <h2>处理摘要</h2>
        <pre>{safe_summary}</pre>
      </section>
    </div>
  </main>
</body>
</html>"""


def _render_llm_section(results: dict[str, Any]) -> str:
    if not results.get("has_draft"):
        return ""

    parts: list[str] = []
    parts.append('<section><h2>LLM 解析结果</h2>')

    tpl_count = results.get("template_count", 0)
    unresolved = results.get("unresolved_count", 0)
    badge_class = "ok" if unresolved == 0 else "warn"
    badge_text = "全部确认" if unresolved == 0 else f"{unresolved} 个待确认"
    parts.append(
        f'<p>项目：{escape(str(results.get("project_name", "")))} | '
        f'主文件：{escape(str(results.get("main_file", "")))} | '
        f'模板数：{tpl_count} | '
        f'<span class="badge {badge_class}">{badge_text}</span></p>')

    templates = results.get("templates", [])
    if templates:
        parts.append('<table class="result-table"><tr><th>模板</th><th>源 Sheet</th><th>目标 Sheet</th><th>映射数</th><th>明细区</th></tr>')
        for tpl in templates:
            has_detail = "是" if tpl.get("has_detail") else "否"
            parts.append(
                f'<tr><td>{escape(str(tpl.get("name", "")))}</td>'
                f'<td>{escape(str(tpl.get("source_sheet", "")))}</td>'
                f'<td>{escape(str(tpl.get("target_sheet", "")))}</td>'
                f'<td>{tpl.get("mapping_count", 0)}</td>'
                f'<td>{has_detail}</td></tr>')
        parts.append("</table>")

    confirm_items = results.get("confirm_items", [])
    if confirm_items:
        parts.append(f'<h3>待确认问题（共 {results.get("confirm_total", 0)} 条）— 请在下方的编辑器中修改草稿来解决</h3>')
        parts.append('<table class="result-table"><tr><th>字段</th><th>当前值</th><th>原因</th></tr>')
        for item in confirm_items:
            parts.append(
                f'<tr><td><code>{escape(str(item.get("field", "")))}</code></td>'
                f'<td>{escape(str(item.get("value", "")))}</td>'
                f'<td>{escape(str(item.get("reason", "")))}</td></tr>')
        parts.append("</table>")

    parts.append(
        '<div class="actions" style="margin-top:14px">'
        '<a class="download secondary" href="/download-llm-draft">下载配置草稿</a>'
        '<a class="download secondary" href="/download-llm-confirm">下载待确认问题</a>'
        '</div>')

    parts.append("</section>")
    return "\n".join(parts)


def _render_editor_section(draft_json: str, validation: dict[str, Any], llm_results: dict[str, Any]) -> str:
    if not draft_json:
        return ""

    parts: list[str] = []
    parts.append('<section><h2>草稿编辑器</h2>')
    parts.append('<p style="font-size:13px;color:var(--muted)">直接修改下方 JSON 来修复待确认项，然后点击<strong>保存并校验</strong>。校验通过后可发布为正式配置。</p>')

    parts.append('<form action="/save-draft" method="post">')
    parts.append(
        f'<textarea class="draft-editor" name="draft_text" spellcheck="false">'
        f'{escape(draft_json)}</textarea>')
    parts.append('<div class="actions" style="margin-top:12px">')
    parts.append('<button type="submit" class="ok-btn">保存并校验</button>')
    parts.append('</div>')
    parts.append('</form>')

    # Promote button (only if validation passed)
    if validation.get("ready"):
        parts.append(
            '<form action="/promote-draft" method="post" style="margin-top:10px">'
            '<button type="submit" class="warn-btn">发布为正式配置</button>'
            '</form>')

    parts.append("</section>")
    return "\n".join(parts)


def _render_validation_section(validation: dict[str, Any]) -> str:
    if not validation:
        return ""

    parts: list[str] = []
    parts.append('<section><h2>校验结果</h2>')

    passed = validation.get("passed", 0)
    need_confirm = validation.get("need_confirm", 0)
    failed = validation.get("failed", 0)
    ready = validation.get("ready", False)
    reason = escape(str(validation.get("reason", "")))

    ready_badge = '<span class="badge ok">可填充</span>' if ready else '<span class="badge warn">不可填充</span>'
    parts.append(f'<p>{ready_badge} — {reason}</p>')

    parts.append('<div class="val-summary">')
    parts.append(f'<div class="val-chip pass">通过 {passed}</div>')
    parts.append(f'<div class="val-chip confirm">待确认 {need_confirm}</div>')
    parts.append(f'<div class="val-chip fail">失败 {failed}</div>')
    parts.append('</div>')

    failed_items = validation.get("failed_items", [])
    if failed_items:
        parts.append('<h3>失败项</h3>')
        parts.append('<table class="val-table"><tr><th>问题</th></tr>')
        for item in failed_items:
            parts.append(f'<tr><td style="color:var(--bad)">{escape(str(item))}</td></tr>')
        parts.append('</table>')

    need_confirm_items = validation.get("need_confirm_items", [])
    if need_confirm_items:
        parts.append('<h3>待确认项</h3>')
        parts.append('<table class="val-table"><tr><th>问题</th></tr>')
        for item in need_confirm_items:
            parts.append(f'<tr><td style="color:var(--warn)">{escape(str(item))}</td></tr>')
        parts.append('</table>')

    parts.append("</section>")
    return "\n".join(parts)


def _render_draft_editor(draft_json: str, validation: dict[str, Any]) -> str:
    """Standalone draft editor page (for /view-draft-json)."""
    safe_json = escape(draft_json)
    val_html = _render_validation_section(validation)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>Draft Editor</title>
  <style>
    body {{ font-family: Arial, "Microsoft YaHei", sans-serif; max-width: 900px; margin: 30px auto; padding: 0 20px; }}
    textarea {{ width: 100%; min-height: 400px; font-family: monospace; font-size: 13px; }}
    button {{ padding: 10px 20px; cursor: pointer; }}
  </style>
</head>
<body>
  <h2>草稿编辑器</h2>
  {val_html}
  <form action="/save-draft" method="post">
    <textarea name="draft_text">{safe_json}</textarea>
    <p><button type="submit">保存并校验</button></p>
  </form>
</body>
</html>"""
